from collections import defaultdict
from datetime import date, datetime, timedelta

from app.models.enums import EstadoLote, TipoAnalisis
from app.models.models import (
    AnalisisLey,
    AnalisisRecuperacion,
    Configuracion,
    Lote,
    MapeoCIP,
    Muestreo,
    Pesaje,
    PruebaMetalurgica,
    SesionDescarga,
)
from app.schemas.dashboard import (
    AcopiadorStats,
    AcopiadorTMH,
    AlertaItem,
    AlertasConfig,
    AlertasResponse,
    AnalisisConteo,
    DashboardKPIs,
    DashboardResponse,
    LoteDashboard,
)
from sqlalchemy.orm import Session

dias_habilitado = 30  # días de almacén para considerar lote "habilitado"


def _calcular_estadoanalisis(
    tiene_cip: bool,
    tiene_ley: bool,
    tiene_rec_completa: bool,
    tiene_humedad: bool,
) -> str:
    if not tiene_cip:
        return "SIN_DATOS"
    if not tiene_humedad:
        return "FALTA_MUESTREO"
    if not tiene_ley:
        return "FALTA_LEY"
    if not tiene_rec_completa:
        return "FALTA_REC"
    return "LISTO"


def _calcular_estado_lote(lote: Lote, dias: int, tiene_recuperacion: bool) -> str:
    if lote.volado:
        return "VOLADO"
    if lote.dirimencia:
        return "DIRIMENCIA"
    if tiene_recuperacion:
        return "COMPLETO"
    if lote.habilitado_ruma or dias >= dias_habilitado:
        return "HABILITADO"
    return "EN_PROCESO"


alert_defaults: dict[str, str] = {
    "alerta_horas_pesado_muestreo": "24",
    "alerta_horas_muestreo_ley": "24",
    "alerta_horas_ley_recuperacion": "72",
    "alerta_dias_volado_stock": "30",
}

alert_desc: dict[str, str] = {
    "alerta_horas_pesado_muestreo": "Horas máx. entre pesaje y primer muestreo",
    "alerta_horas_muestreo_ley": "Horas máx. entre muestreo y análisis de ley",
    "alerta_horas_ley_recuperacion": "Horas máx. entre inicio de pruebas y recuperación",
    "alerta_dias_volado_stock": "Días máx. que un lote volado puede estar sin ruma",
}


def _dt(val) -> datetime:
    """Normaliza date/datetime a datetime."""
    if isinstance(val, datetime):
        return val
    return datetime.combine(val, datetime.min.time())


def _severidad(delta: timedelta, umbral: timedelta) -> str:
    ratio = delta.total_seconds() / max(umbral.total_seconds(), 1)
    if ratio >= 2.0:
        return "CRITICA"
    if ratio >= 1.5:
        return "ALTA"
    return "MEDIA"


def _make_alerta(
    tipo: str,
    delta: timedelta,
    umbral: timedelta,
    desc: str,
    fecha: datetime,
    ip: str,
    proveedor: str,
    acopiador: str,
) -> AlertaItem:
    return AlertaItem(
        tipo=tipo,
        severidad=_severidad(delta, umbral),
        ip=ip,
        proveedor=proveedor,
        acopiador=acopiador,
        horas_retraso=round(delta.total_seconds() / 3600, 1),
        descripcion=desc,
        fecha_ref=fecha,
    )


def _cargar_config_alertas(db: Session) -> AlertasConfig:
    rows = (
        db.query(Configuracion.clave, Configuracion.valor)
        .filter(Configuracion.clave.in_(alert_defaults.keys()))
        .all()
    )
    vals = {k: v for k, v in rows}
    return AlertasConfig(
        horas_pesado_muestreo=float(vals.get("alerta_horas_pesado_muestreo", 24)),
        horas_muestreo_ley=float(vals.get("alerta_horas_muestreo_ley", 24)),
        horas_ley_recuperacion=float(vals.get("alerta_horas_ley_recuperacion", 72)),
        dias_volado_stock=float(vals.get("alerta_dias_volado_stock", 30)),
    )


def obtener_resumen_dashboard(db: Session) -> DashboardResponse:
    lotes_db = db.query(Lote).filter(~Lote.eliminado).all()

    # Pre-cargar sets en queries únicas (evita N+1)
    ids_con_cip: set[int] = {row[0] for row in db.query(MapeoCIP.lote_id).distinct().all()}

    # Pre-cargar todas las Leyes vigentes (excluyendo minero)
    leyes_db = (
        db.query(
            AnalisisLey.lote_id,
            AnalisisLey.tipo_analisis,
            AnalisisLey.ley_final,
            AnalisisLey.ley_gr_tm,
        )
        .filter(
            AnalisisLey.vigente == True,  # noqa: E712
            AnalisisLey.tipo_analisis != TipoAnalisis.MINERO,
            AnalisisLey.material == "Au",
        )
        .all()
    )

    leyes_por_lote = defaultdict(list)
    ids_con_ley = set()
    for lote_id, tipo, ley_final, ley_gr_tm in leyes_db:
        leyes_por_lote[lote_id].append(
            {
                "tipo": tipo,
                "valor": float(ley_final) if ley_final is not None else None,  # oz/TC → ley_avg
                "valor_gr_tm": float(ley_gr_tm) if ley_gr_tm is not None else None,  # gr/TM → KPIs
            }
        )

    # Pre-cargar todas las Recuperaciones vigentes
    recs_db = (
        db.query(
            AnalisisRecuperacion.lote_id,
            AnalisisRecuperacion.estado,
            AnalisisRecuperacion.recuperacion,
        )
        .filter(AnalisisRecuperacion.vigente == True)  # noqa: E712
        .all()
    )

    recs_por_lote = defaultdict(list)
    ids_con_recuperacion = set()
    for lote_id, estado, valor in recs_db:
        ids_con_recuperacion.add(lote_id)
        recs_por_lote[lote_id].append({"estado": estado, "valor": valor})

    # ids con recuperación COMPLETADA y sin ninguna PENDIENTE
    ids_rec_completa: set[int] = set()
    ids_rec_pendiente: set[int] = set()
    for lote_id, recs in recs_por_lote.items():
        has_completado = any(r["estado"] in ("COMPLETADO", "CERT_COMERCIAL") for r in recs)
        has_pendiente = any(r["estado"] == "PENDIENTE" for r in recs)
        if has_pendiente:
            ids_rec_pendiente.add(lote_id)
        if has_completado and not has_pendiente:
            ids_rec_completa.add(lote_id)

    kpis = DashboardKPIs()
    lotes_resumen = []
    analisis_counts: dict[str, int] = {}
    stats_acop: dict[str, dict] = {}

    # Estructura para almacenar sumatorias: { acopiador: { mes_num: sum_tmh } }
    tmh_por_acopiador_mes = defaultdict(lambda: defaultdict(float))

    for lote in lotes_db:
        pesaje = db.query(Pesaje).filter(Pesaje.lote_id == lote.id).first()
        tmh = 0.0
        if pesaje and pesaje.peso_inicial and pesaje.peso_final:
            if pesaje.peso_inicial > pesaje.peso_final:
                tmh = float(pesaje.peso_inicial - pesaje.peso_final)

        kpis.tmh_stock += tmh

        h2o_porc = None
        tms = None

        muestreo = (
            db.query(Muestreo)
            .filter(Muestreo.lote_id == lote.id)
            .order_by(Muestreo.creado_en.desc())
            .first()
        )

        if muestreo and muestreo.peso_humedo and muestreo.peso_seco:
            ph = float(muestreo.peso_humedo)
            ps = float(muestreo.peso_seco)
            if ph > 0:
                h2o_porc = round(((ph - ps) / ph) * 100, 2)
                tms = round(tmh * (1 - (h2o_porc / 100)), 3)
                kpis.tms_stock += tms

        # tiene_humedad DEBE evaluarse después del bloque de muestreo
        tiene_humedad = tms is not None

        sesion = db.query(SesionDescarga).filter(SesionDescarga.id == lote.sesion_id).first()
        proveedor_nombre = "---"
        proveedor_ruc = "---"
        acopiador_nombre = "---"

        if sesion and getattr(sesion, "provacop", None):
            if getattr(sesion.provacop, "proveedor", None):
                proveedor_nombre = sesion.provacop.proveedor.razon_social
                proveedor_ruc = sesion.provacop.proveedor.ruc
            if getattr(sesion.provacop, "acopiador", None):
                acopiador_nombre = sesion.provacop.acopiador.razon_social

        # Días en almacén
        fecha_rec = None
        if lote.pesajes:
            dt = lote.pesajes[0].fecha_fin
            fecha_rec = dt.date() if hasattr(dt, "date") else dt
        dias = (date.today() - fecha_rec).days if fecha_rec else 0

        is_volado_30d = bool(lote.volado) and dias >= dias_habilitado
        is_facturado_o_pagado = lote.estado in (EstadoLote.FACTURADO, EstadoLote.PAGADO)
        habilitado_ruma = bool(lote.habilitado_ruma) or is_volado_30d or is_facturado_o_pagado

        # --- ACUMULACIÓN DE TMH POR MES Y ACOPIADOR ---
        if acopiador_nombre and fecha_rec:
            mes_num = fecha_rec.month
            tmh_por_acopiador_mes[acopiador_nombre][mes_num] += tmh

        estadoanalisis = _calcular_estadoanalisis(
            tiene_cip=lote.id in ids_con_cip,
            tiene_ley=lote.id in ids_con_ley,
            tiene_rec_completa=lote.id in ids_rec_completa,
            tiene_humedad=tiene_humedad,
        )
        analisis_counts[estadoanalisis] = analisis_counts.get(estadoanalisis, 0) + 1

        # --- CÁLCULO DE RECUPERACIÓN PROMEDIO O COMERCIAL ---
        rec_prom = None
        recs_del_lote = recs_por_lote.get(lote.id, [])
        if recs_del_lote:
            # Buscar primero si tiene certificado comercial con valor registrado
            rec_comercial = next(
                (
                    r["valor"]
                    for r in recs_del_lote
                    if r["estado"] == "CERT_COMERCIAL" and r["valor"] is not None
                ),
                None,
            )

            if rec_comercial is not None:
                rec_prom = float(rec_comercial)
            else:
                # Si no hay comercial, se promedian las que tengan estado "COMPLETADO"
                valores_rec = [
                    float(r["valor"])
                    for r in recs_del_lote
                    if r["estado"] == "COMPLETADO" and r["valor"] is not None
                ]
                if valores_rec:
                    rec_prom = round(sum(valores_rec) / len(valores_rec), 2)

        # --- CÁLCULO DE LEY PROMEDIO O COMERCIAL ---
        ley_prom = None  # oz/TC → va a ley_avg en LoteDashboard
        ley_gr_tm_prom = None  # gr/TM → se usa para acumular KPI de oro
        leyes_del_lote = leyes_por_lote.get(lote.id, [])
        if leyes_del_lote:
            # Preferir ley comercial
            comercial = next(
                (
                    ley
                    for ley in leyes_del_lote
                    if ley["tipo"] in (TipoAnalisis.COMERCIAL, "ley_comercial")
                    and ley["valor"] is not None
                ),
                None,
            )
            if comercial is not None:
                ley_prom = float(comercial["valor"])  # oz/TC
                ley_gr_tm_prom = float(comercial["valor_gr_tm"] or 0) or ley_prom * 34.2857
            else:
                # Promedio de todas las leyes válidas
                valores_oz = [
                    float(ley["valor"]) for ley in leyes_del_lote if ley["valor"] is not None
                ]
                valores_gr = [
                    float(ley["valor_gr_tm"])
                    for ley in leyes_del_lote
                    if ley.get("valor_gr_tm") is not None
                ]
                if valores_oz:
                    ley_prom = round(sum(valores_oz) / len(valores_oz), 4)  # oz/TC
                if valores_gr:
                    ley_gr_tm_prom = round(sum(valores_gr) / len(valores_gr), 3)  # gr/TM
                elif ley_prom is not None:
                    ley_gr_tm_prom = ley_prom * 34.2857  # conversión de respaldo

        # KPI oro: usa gr/TM (TMS × ley_gr_tm = gramos)
        if tms is not None and ley_gr_tm_prom is not None:
            au_lote = tms * ley_gr_tm_prom
            kpis.au_real_100 += au_lote
            if rec_prom is not None:
                kpis.au_real_rec += au_lote * rec_prom / 100

            # oz_habilitados: solo lotes habilitados por flag/volado, NO por ya estar facturados
            habilitado_para_kpi = (
                bool(lote.habilitado_ruma) or is_volado_30d
            ) and not is_facturado_o_pagado
            if habilitado_para_kpi and lote.ruma_id is None:
                kpis.oz_habilitados += au_lote / 31.1035

        # Stats por acopiador (acumular fuera del bloque de ley para contar lotes siempre)
        if acopiador_nombre and acopiador_nombre != "---":
            s = stats_acop.setdefault(
                acopiador_nombre, {"lotes": 0, "tms": 0.0, "oz_sum": 0.0, "ley_tms": 0.0}
            )
            s["lotes"] += 1
            if tms:
                s["tms"] += tms
            if tms and ley_gr_tm_prom:
                s["oz_sum"] += (tms * ley_gr_tm_prom) / 31.1035
                s["ley_tms"] += tms * ley_gr_tm_prom

        lotes_resumen.append(
            LoteDashboard(
                ip=lote.ip,
                tmh=round(tmh, 3),
                tms=tms,
                h2o_porc=h2o_porc,
                proveedor=proveedor_nombre,
                ruc=proveedor_ruc,
                ley_avg=ley_prom,
                rec_porc=rec_prom,
                acopiador=acopiador_nombre,
                estado=lote.estado if lote.estado else "RECEPCIONADO",
                estado_analisis=estadoanalisis,
                habilitado_ruma=habilitado_ruma,
                volado=bool(lote.volado),
                dirimencia=bool(lote.dirimencia),
                dias_almacen=dias,
                tiene_rec_pendiente=lote.id in ids_rec_pendiente,
            )
        )

    kpis.tmh_stock = round(kpis.tmh_stock, 2)
    kpis.tms_stock = round(kpis.tms_stock, 2)
    kpis.au_real_100 = round(kpis.au_real_100, 2)
    kpis.au_real_rec = round(kpis.au_real_rec, 2)
    kpis.oz_stock = round(kpis.au_real_100 / 31.1035, 3)
    kpis.oz_habilitados = round(kpis.oz_habilitados, 3)

    lotes_resumen.sort(key=lambda x: x.ip, reverse=True)

    acopiadores_tmh_list = []
    for acop_name, meses_dict in tmh_por_acopiador_mes.items():
        acopiadores_tmh_list.append(
            AcopiadorTMH(
                acopiador=acop_name,
                enero=round(meses_dict[1], 2),
                febrero=round(meses_dict[2], 2),
                marzo=round(meses_dict[3], 2),
                abril=round(meses_dict[4], 2),
                mayo=round(meses_dict[5], 2),
                junio=round(meses_dict[6], 2),
                julio=round(meses_dict[7], 2),
                agosto=round(meses_dict[8], 2),
                septiembre=round(meses_dict[9], 2),
                octubre=round(meses_dict[10], 2),
                noviembre=round(meses_dict[11], 2),
                diciembre=round(meses_dict[12], 2),
                total=round(sum(meses_dict.values()), 2),
            )
        )
    acopiadores_tmh_list.sort(key=lambda x: x.total, reverse=True)

    acopiadores_stats_list: list[AcopiadorStats] = []
    for name, s in stats_acop.items():
        ley_p = round(s["ley_tms"] / s["tms"], 3) if s["tms"] > 0 else None
        acopiadores_stats_list.append(
            AcopiadorStats(
                acopiador=name,
                lotes=s["lotes"],
                tms=round(s["tms"], 2),
                oz=round(s["oz_sum"], 3),
                ley_prom=ley_p,
            )
        )
    acopiadores_stats_list.sort(key=lambda x: x.oz, reverse=True)

    conteo = AnalisisConteo(
        listo=analisis_counts.get("LISTO", 0),
        falta_rec=analisis_counts.get("FALTA_REC", 0),
        falta_ley=analisis_counts.get("FALTA_LEY", 0),
        falta_muestreo=analisis_counts.get("FALTA_MUESTREO", 0),
        sin_datos=analisis_counts.get("SIN_DATOS", 0),
    )

    return DashboardResponse(
        kpis=kpis,
        lotes=lotes_resumen,
        acopiadores_tmh=acopiadores_tmh_list,
        analisis_conteo=conteo,
        acopiadores_stats=acopiadores_stats_list,
    )


def generar_excel_dashboard(data: DashboardResponse, tipo: str, clave: str):
    import io

    import msoffcrypto
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    gold_fill = PatternFill("solid", fgColor="1C1A09")
    gold_font = Font(bold=True, color="C9A227")
    center = Alignment(horizontal="center")

    def _autowidth(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 45)

    def _header_row(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = gold_fill
            c.font = gold_font
            c.alignment = center

    analisis = {
        "SIN_DATOS": "Sin datos",
        "FALTA_MUESTREO": "Falta humedad",
        "FALTA_LEY": "Falta ley",
        "FALTA_REC": "Falta rec.",
        "LISTO": "Listo",
    }

    wb = Workbook()

    if tipo == "lotes":
        ws = wb.active
        ws.title = "Lotes"
        _header_row(
            ws,
            [
                "IP",
                "TMH",
                "TMS",
                "%H2O",
                "Proveedor",
                "RUC",
                "Ley gr/TM",
                "% Rec",
                "Acopiador",
                "Estado",
                "Análisis",
                "Días Almacén",
                "Habilitado Ruma",
                "Volado",
            ],
        )
        for lt in data.lotes:
            ws.append(
                [
                    lt.ip,
                    lt.tmh,
                    lt.tms,
                    lt.h2o_porc,
                    lt.proveedor,
                    lt.ruc,
                    lt.ley_avg,
                    lt.rec_porc,
                    lt.acopiador,
                    lt.estado,
                    analisis.get(lt.estado_analisis, lt.estado_analisis),
                    lt.dias_almacen,
                    "Sí" if lt.habilitado_ruma else "No",
                    "Sí" if lt.volado else "No",
                ]
            )
    else:
        ws = wb.active
        ws.title = "Acopiadores TMH"
        _header_row(
            ws,
            [
                "Acopiador",
                "Ene",
                "Feb",
                "Mar",
                "Abr",
                "May",
                "Jun",
                "Jul",
                "Ago",
                "Set",
                "Oct",
                "Nov",
                "Dic",
                "Total",
            ],
        )
        for a in data.acopiadores_tmh:
            ws.append(
                [
                    a.acopiador,
                    a.enero,
                    a.febrero,
                    a.marzo,
                    a.abril,
                    a.mayo,
                    a.junio,
                    a.julio,
                    a.agosto,
                    a.septiembre,
                    a.octubre,
                    a.noviembre,
                    a.diciembre,
                    a.total,
                ]
            )
        ws2 = wb.create_sheet("Stats por Acopiador")
        _header_row(ws2, ["Acopiador", "Lotes", "TMS", "Oz en Stock", "Ley Prom gr/TM"])
        for s in data.acopiadores_stats:
            ws2.append([s.acopiador, s.lotes, s.tms, s.oz, s.ley_prom])

    _autowidth(wb.active)

    raw = io.BytesIO()
    wb.save(raw)
    raw.seek(0)

    encrypted = io.BytesIO()
    msoffcrypto.OfficeFile(raw).encrypt(clave, encrypted)
    encrypted.seek(0)
    return encrypted


# Alertas:
def obtener_alertas(db: Session) -> AlertasResponse:
    cfg = _cargar_config_alertas(db)
    now = datetime.utcnow()
    alertas: list[AlertaItem] = []

    # ── Lotes activos ────────────────────────────────────────────────
    lotes = db.query(Lote).filter(~Lote.eliminado, Lote.estado != EstadoLote.PAGADO).all()
    ids = [lt.id for lt in lotes]
    if not ids:
        return AlertasResponse(alertas=[], config=cfg)

    # ── 1. Nombres y Sesiones ──
    sesiones = {
        s.id: s
        for s in db.query(SesionDescarga)
        .filter(SesionDescarga.id.in_({lt.sesion_id for lt in lotes}))
        .all()
    }

    def _nombres(lote: Lote) -> tuple[str, str]:
        s = sesiones.get(lote.sesion_id)
        prov = acop = "---"
        if s and getattr(s, "provacop", None):
            if getattr(s.provacop, "proveedor", None):
                prov = s.provacop.proveedor.razon_social
            if getattr(s.provacop, "acopiador", None):
                acop = s.provacop.acopiador.razon_social
        return prov, acop

    # ── 2. Extracción de Fechas (Alineado 100% con el Dashboard) ──

    # Pesaje: Tomamos el PRIMER registro (id.asc) al igual que hace lote.pesajes[0]
    pesajes = (
        db.query(Pesaje.lote_id, Pesaje.fecha_fin)
        .filter(Pesaje.lote_id.in_(ids))
        .order_by(Pesaje.id.asc())
        .all()
    )
    pesaje_fecha: dict[int, datetime] = {}
    for pid, pfecha in pesajes:
        if pfecha and pid not in pesaje_fecha:
            pesaje_fecha[pid] = _dt(pfecha)

    # Muestreo: Tomamos el ÚLTIMO registro válido
    muestreos = (
        db.query(Muestreo)
        .filter(Muestreo.lote_id.in_(ids))
        .order_by(Muestreo.creado_en.asc())
        .all()
    )
    muestreo_fecha: dict[int, datetime] = {}
    tiene_muestreo: dict[int, bool] = {}
    for m in muestreos:
        # Regla idéntica al dashboard: debe tener ambos pesos mayores a 0
        if m.peso_humedo and m.peso_seco and float(m.peso_humedo) > 0:
            muestreo_fecha[m.lote_id] = _dt(m.creado_en)
            tiene_muestreo[m.lote_id] = True
        else:
            tiene_muestreo[m.lote_id] = False

    # Ley: Tomamos el ÚLTIMO registro válido (Corrección del typo: tipo_analisis)
    leyes = (
        db.query(AnalisisLey)
        .filter(
            AnalisisLey.lote_id.in_(ids),
            AnalisisLey.vigente,
            AnalisisLey.tipo_analisis != TipoAnalisis.MINERO,
        )
        .order_by(AnalisisLey.creado_en.asc())
        .all()
    )

    ley_fecha: dict[int, datetime] = {}
    tiene_ley: dict[int, bool] = {}
    for le in leyes:
        if le.ley_gr_tm is not None:
            ley_fecha[le.lote_id] = _dt(le.creado_en)
            tiene_ley[le.lote_id] = True
        else:
            tiene_ley[le.lote_id] = False

    # Recuperación: Completada y sin pendientes
    recs = (
        db.query(AnalisisRecuperacion)
        .filter(AnalisisRecuperacion.lote_id.in_(ids), AnalisisRecuperacion.vigente)
        .all()
    )

    recs_dict = defaultdict(list)
    for r in recs:
        recs_dict[r.lote_id].append(r.estado)

    ids_con_rec = {
        lid
        for lid, estados in recs_dict.items()
        if any(e in ("COMPLETADO", "CERT_COMERCIAL") for e in estados)
        and not any(e == "PENDIENTE" for e in estados)
    }

    # Pruebas Metalúrgicas
    pruebas = (
        db.query(PruebaMetalurgica.lote_id, PruebaMetalurgica.fecha_ingreso)
        .filter(PruebaMetalurgica.lote_id.in_(ids))
        .all()
    )
    prueba_fecha = {}
    for pid, pfecha in pruebas:
        if pfecha:
            dt_p = _dt(pfecha)
            if pid not in prueba_fecha or dt_p > prueba_fecha[pid]:
                prueba_fecha[pid] = dt_p

    # ── 3. Evaluación de Alertas (Cascada - Cuello de Botella) ──
    umb_muestreo = timedelta(hours=cfg.horas_pesado_muestreo)
    umb_ley = timedelta(hours=cfg.horas_muestreo_ley)
    umb_rec = timedelta(hours=cfg.horas_ley_recuperacion)
    umb_volado = timedelta(days=cfg.dias_volado_stock)

    for lote in lotes:
        ip, prov, acop = lote.ip, *_nombres(lote)

        fp = pesaje_fecha.get(lote.id)
        fm = muestreo_fecha.get(lote.id)
        fl = ley_fecha.get(lote.id)
        fr = prueba_fecha.get(lote.id) or fl

        # Alerta aislada: Volado sin ruma
        if lote.volado and lote.ruma_id is None and fp:
            delta = now - fp
            if delta >= umb_volado:
                alertas.append(
                    _make_alerta(
                        "VOLADO_STOCK",
                        delta,
                        umb_volado,
                        f"Lote volado sin asignar a ruma hace {int(delta.days)} días",
                        fp,
                        ip,
                        prov,
                        acop,
                    )
                )

        # Alertas en cascada (elif) - Bloquea si la etapa actual no está completada
        if fp and not tiene_muestreo.get(lote.id, False):
            delta = now - fp
            if delta >= umb_muestreo:
                alertas.append(
                    _make_alerta(
                        "RETRASO_MUESTREO",
                        delta,
                        umb_muestreo,
                        f"Sin muestreo {delta.total_seconds()/3600:.1f}h después del pesaje",
                        fp,
                        ip,
                        prov,
                        acop,
                    )
                )

        elif fm and not tiene_ley.get(lote.id, False):
            delta = now - fm
            if delta >= umb_ley:
                alertas.append(
                    _make_alerta(
                        "RETRASO_LEY",
                        delta,
                        umb_ley,
                        f"Sin análisis de ley {delta.total_seconds()/3600:.1f}h después del muestreo",
                        fm,
                        ip,
                        prov,
                        acop,
                    )
                )

        elif fl and lote.id not in ids_con_rec and fr:
            delta = now - fr
            if delta >= umb_rec:
                origen = "inicio de pruebas" if lote.id in prueba_fecha else "análisis de ley"
                alertas.append(
                    _make_alerta(
                        "RETRASO_RECUPERACION",
                        delta,
                        umb_rec,
                        f"Sin recuperación {delta.total_seconds()/3600:.1f}h después del {origen}",
                        fr,
                        ip,
                        prov,
                        acop,
                    )
                )

    # Retorno de resultados
    return AlertasResponse(
        alertas=alertas,
        config=cfg,
        total_criticas=sum(1 for a in alertas if a.severidad == "CRITICA"),
        total_altas=sum(1 for a in alertas if a.severidad == "ALTA"),
        total_medias=sum(1 for a in alertas if a.severidad == "MEDIA"),
    )


def actualizar_config_alertas(db: Session, config: AlertasConfig) -> None:
    updates = {
        "alerta_horas_pesado_muestreo": str(config.horas_pesado_muestreo),
        "alerta_horas_muestreo_ley": str(config.horas_muestreo_ley),
        "alerta_horas_ley_recuperacion": str(config.horas_ley_recuperacion),
        "alerta_dias_volado_stock": str(config.dias_volado_stock),
    }
    for clave, valor in updates.items():
        row = db.query(Configuracion).filter(Configuracion.clave == clave).first()
        if row:
            row.valor = valor
        else:
            db.add(Configuracion(clave=clave, valor=valor, descripcion=alert_desc.get(clave, "")))
    db.commit()


# =============================================================================
# TRAZABILIDAD POR LOTE
# =============================================================================

from app.models.models import (  # noqa: E402
    Liquidacion,
    LiquidacionLote,
    ProveedorAcopiador,
    Ruma,
    RumaCampana,
    Usuario,
)
from app.schemas.dashboard import (  # noqa: E402
    AccionRegistro,
    TrazabilidadAnalisisLey,
    TrazabilidadAnalisisRec,
    TrazabilidadAuditoria,
    TrazabilidadLiquidacion,
    TrazabilidadLoteResponse,
    TrazabilidadMuestreo,
    TrazabilidadPesaje,
    TrazabilidadPrueba,
    TrazabilidadRuma,
    TrazabilidadSesion,
    UsuarioResumen,
)
from fastapi import HTTPException  # noqa: E402
from sqlalchemy.orm import joinedload, selectinload  # noqa: E402


def _usuario_res(u: "Usuario | None") -> "UsuarioResumen | None":
    """Convierte un objeto Usuario en UsuarioResumen (o None)."""
    if u is None:
        return None
    return UsuarioResumen(
        id=u.id,
        nombre_completo=u.nombre_completo,
        rol=u.rol.codigo if u.rol else "",
    )


def _accion(u_map: dict, user_id: "int | None", fecha) -> AccionRegistro:
    """Construye AccionRegistro buscando el usuario en el mapa batched."""
    usuario = u_map.get(user_id) if user_id else None
    return AccionRegistro(
        por=_usuario_res(usuario),
        fecha=fecha,
    )


def obtener_trazabilidad_lote(db: Session, ip: str) -> TrazabilidadLoteResponse:
    """
    Retorna la traza administrativa completa de un lote identificado por su IP.
    Diseñado sin N+1: una query principal + una sola query de usuarios batch.
    """
    lote = (
        db.query(Lote)
        .options(
            # Sesión + provacop + proveedor + acopiador
            joinedload(Lote.sesion)
            .joinedload(SesionDescarga.provacop)
            .joinedload(ProveedorAcopiador.proveedor),
            joinedload(Lote.sesion)
            .joinedload(SesionDescarga.provacop)
            .joinedload(ProveedorAcopiador.acopiador),
            # Pesajes
            selectinload(Lote.pesajes),
            # Muestreos
            selectinload(Lote.muestreos),
            # Prueba metalúrgica
            selectinload(Lote.prueba_metalurgica),
            # Análisis ley + descartador
            selectinload(Lote.analisis_ley)
            .joinedload(AnalisisLey.descartador)
            .joinedload(Usuario.rol),
            # Análisis recuperación + descartador
            selectinload(Lote.analisis_recuperacion)
            .joinedload(AnalisisRecuperacion.descartador)
            .joinedload(Usuario.rol),
            # Ruma → rumas_campana → campaña
            joinedload(Lote.ruma).selectinload(Ruma.rumas_campana).joinedload(RumaCampana.campana),
            # Liquidación → lote pivot → liquidacion → cerrador + rol
            selectinload(Lote.liquidaciones_lotes)
            .joinedload(LiquidacionLote.liquidacion)
            .joinedload(Liquidacion.cerrador)
            .joinedload(Usuario.rol),
            # Usuario habilitación + modificador estado
            joinedload(Lote.usuario_habilitacion).joinedload(Usuario.rol),
            joinedload(Lote.modificador_estado).joinedload(Usuario.rol),
        )
        .filter(Lote.ip == ip, Lote.eliminado == False)  # noqa: E712
        .first()
    )
    if not lote:
        raise HTTPException(status_code=404, detail=f"Lote con IP '{ip}' no encontrado")

    # ── Recolectar todos los creado_por que necesitan resolverse via batch ──
    ids_usuarios: set[int] = set()

    def _add(uid):
        if uid:
            ids_usuarios.add(uid)

    _add(lote.sesion.creado_por if lote.sesion else None)
    _add(lote.creado_por)
    _add(lote.habilitado_por)
    _add(lote.estado_modificado_por)

    for p in lote.pesajes:
        _add(p.creado_por)
    for m in lote.muestreos:
        _add(m.creado_por)
    if lote.prueba_metalurgica:
        pm = (
            lote.prueba_metalurgica[0]
            if isinstance(lote.prueba_metalurgica, list)
            else lote.prueba_metalurgica
        )
        _add(pm.creado_por)
    for a in lote.analisis_ley:
        _add(a.creado_por)
    for a in lote.analisis_recuperacion:
        _add(a.creado_por)
    for ll in lote.liquidaciones_lotes:
        if ll.liquidacion:
            _add(ll.liquidacion.creado_por)

    u_map: dict[int, Usuario] = (
        {
            u.id: u
            for u in db.query(Usuario)
            .filter(Usuario.id.in_(ids_usuarios))
            .options(joinedload(Usuario.rol))
            .all()
        }
        if ids_usuarios
        else {}
    )

    # ── Sesión ──
    s = lote.sesion
    proveedor_nombre = "---"
    proveedor_ruc = None
    acopiador_nombre = None
    if s and getattr(s, "provacop", None):
        if getattr(s.provacop, "proveedor", None):
            proveedor_nombre = s.provacop.proveedor.razon_social
            proveedor_ruc = s.provacop.proveedor.ruc
        if getattr(s.provacop, "acopiador", None):
            acopiador_nombre = s.provacop.acopiador.razon_social

    sesion_out = TrazabilidadSesion(
        id=s.id,
        placa=s.placa,
        carreta=s.carreta,
        conductor=s.conductor,
        transportista=s.transportista,
        guia_remision=s.guia_remision,
        guia_transporte=s.guia_transporte,
        registro=_accion(u_map, s.creado_por, s.creado_en),
    )

    # ── Pesajes ──
    pesajes_out = [
        TrazabilidadPesaje(
            numero_ticket=p.numero_ticket,
            sacos=p.sacos,
            granel=bool(p.granel),
            peso_inicial=float(p.peso_inicial),
            peso_final=float(p.peso_final),
            peso_neto=float(p.peso_neto),
            fecha_inicio=p.fecha_inicio,
            fecha_fin=p.fecha_fin,
            es_manual=bool(p.es_manual),
            justificacion_manual=p.justificacion_manual,
            registro=_accion(u_map, p.creado_por, p.creado_en),
        )
        for p in sorted(lote.pesajes, key=lambda x: x.id)
    ]

    # ── Muestreos ──
    muestreos_out = [
        TrazabilidadMuestreo(
            intento=m.intento,
            peso_humedo=float(m.peso_humedo),
            peso_seco=float(m.peso_seco),
            porcentaje_humedad=float(m.porcentaje_humedad)
            if m.porcentaje_humedad is not None
            else None,
            tms_calculado=float(m.tms_calculado) if m.tms_calculado is not None else None,
            observaciones=m.observaciones,
            registro=_accion(u_map, m.creado_por, m.creado_en),
        )
        for m in sorted(lote.muestreos, key=lambda x: x.intento)
    ]

    # ── Prueba metalúrgica ──
    prueba_out = None
    prueba_raw = lote.prueba_metalurgica
    if prueba_raw:
        pm = prueba_raw[0] if isinstance(prueba_raw, list) else prueba_raw
        from datetime import timedelta

        fecha_salida = (pm.fecha_ingreso + timedelta(hours=48)) if pm.fecha_ingreso else None
        prueba_out = TrazabilidadPrueba(
            cip=pm.cip,
            fecha_ingreso=pm.fecha_ingreso,
            fecha_salida=fecha_salida,
            malla_porcentaje=float(pm.malla_porcentaje)
            if pm.malla_porcentaje is not None
            else None,
            porcentaje_nacn=float(pm.porcentaje_nacn) if pm.porcentaje_nacn is not None else None,
            ph_inicial=float(pm.ph_inicial) if pm.ph_inicial is not None else None,
            ph_final=float(pm.ph_final) if pm.ph_final is not None else None,
            adicion_nacn=float(pm.adicion_nacn) if pm.adicion_nacn is not None else None,
            adicion_naoh=float(pm.adicion_naoh) if pm.adicion_naoh is not None else None,
            gasto_agno3=float(pm.gasto_agno3) if pm.gasto_agno3 is not None else None,
            registro=_accion(u_map, pm.creado_por, pm.creado_en),
        )

    # ── Análisis de ley ──
    analisis_ley_out = [
        TrazabilidadAnalisisLey(
            id=a.id,
            cip=a.cip,
            laboratorio=a.laboratorio,
            tipo_analisis=a.tipo_analisis,
            material=a.material or "Au",
            ley_final=float(a.ley_final) if a.ley_final is not None else None,
            ley_gr_tm=float(a.ley_gr_tm) if a.ley_gr_tm is not None else None,
            origen_datos=a.origen_datos or "MANUAL",
            fecha_analisis=a.fecha_analisis,
            certificado_url=a.certificado_url,
            vigente=bool(a.vigente),
            descarte=AccionRegistro(
                por=_usuario_res(a.descartador),
                fecha=a.fecha_descarte,
            )
            if not a.vigente and a.descartado_por
            else None,
            justificacion_descarte=a.justificacion_descarte,
            registro=_accion(u_map, a.creado_por, a.creado_en),
        )
        for a in sorted(lote.analisis_ley, key=lambda x: x.id)
        if not getattr(a, "eliminado", False)
    ]

    # ── Análisis de recuperación ──
    analisis_rec_out = [
        TrazabilidadAnalisisRec(
            id=a.id,
            cip=a.cip,
            laboratorio=a.laboratorio,
            ley_cabeza=float(a.ley_cabeza) if a.ley_cabeza is not None else None,
            ley_cola=float(a.ley_cola) if a.ley_cola is not None else None,
            ley_liquido=float(a.ley_liquido) if a.ley_liquido is not None else None,
            recuperacion=float(a.recuperacion) if a.recuperacion is not None else None,
            estado=a.estado,
            origen_datos=a.origen_datos or "MANUAL",
            fecha_analisis=a.fecha_analisis,
            certificado_url=a.certificado_url,
            vigente=bool(a.vigente),
            descarte=AccionRegistro(
                por=_usuario_res(a.descartador),
                fecha=a.fecha_descarte,
            )
            if not a.vigente and a.descartado_por
            else None,
            justificacion_descarte=a.justificacion_descarte,
            registro=_accion(u_map, a.creado_por, a.creado_en),
        )
        for a in sorted(lote.analisis_recuperacion, key=lambda x: x.id)
        if not getattr(a, "eliminado", False)
    ]

    # ── Ruma ──
    ruma_out = None
    if lote.ruma:
        r = lote.ruma
        campana_codigo = None
        if r.rumas_campana:
            rc = r.rumas_campana[0]
            if rc.campana:
                campana_codigo = rc.campana.codigo
        ruma_out = TrazabilidadRuma(
            codigo=r.codigo,
            estado=r.estado,
            fecha_creacion=r.fecha_creacion,
            campana=campana_codigo,
        )

    # ── Liquidación (primera si existe) ──
    liquidacion_out = None
    if lote.liquidaciones_lotes:
        ll = lote.liquidaciones_lotes[0]
        liq = ll.liquidacion
        if liq:
            cierre = None
            if liq.cerrado_por and liq.cerrador:
                cierre = AccionRegistro(
                    por=_usuario_res(liq.cerrador),
                    fecha=liq.fecha_cierre,
                )
            liquidacion_out = TrazabilidadLiquidacion(
                id=liq.id,
                numero_liquidacion=liq.numero_liquidacion,
                estado=liq.estado,
                precio_oro_usd=float(liq.precio_oro_usd)
                if liq.precio_oro_usd is not None
                else None,
                valor_total_usd=float(liq.valor_total_usd)
                if liq.valor_total_usd is not None
                else None,
                fino_recuperable=float(ll.fino_recuperable)
                if ll.fino_recuperable is not None
                else None,
                ley_comercial=float(ll.ley_comercial) if ll.ley_comercial is not None else None,
                usa_dirimencia=bool(ll.usa_dirimencia),
                generacion=_accion(u_map, liq.creado_por, liq.creado_en),
                cierre=cierre,
            )

    # ── Auditoría ──
    auditoria_out = TrazabilidadAuditoria(
        registro_lote=_accion(u_map, lote.creado_por, lote.creado_en),
        habilitacion_ruma=AccionRegistro(
            por=_usuario_res(lote.usuario_habilitacion),
            fecha=lote.fecha_habilitacion,
        ),
        cambio_estado=AccionRegistro(
            por=_usuario_res(lote.modificador_estado),
            fecha=lote.fecha_modificacion_estado,
        ),
    )

    return TrazabilidadLoteResponse(
        ip=lote.ip,
        estado=lote.estado or "RECEPCIONADO",
        tipo_material=lote.tipo_material,
        volado=bool(lote.volado),
        dirimencia=bool(lote.dirimencia),
        habilitado_ruma=bool(lote.habilitado_ruma),
        proveedor=proveedor_nombre,
        ruc_proveedor=proveedor_ruc,
        acopiador=acopiador_nombre,
        sesion=sesion_out,
        pesajes=pesajes_out,
        muestreos=muestreos_out,
        prueba_metalurgica=prueba_out,
        analisis_ley=analisis_ley_out,
        analisis_recuperacion=analisis_rec_out,
        ruma=ruma_out,
        liquidacion=liquidacion_out,
        auditoria=auditoria_out,
    )
