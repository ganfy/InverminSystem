"""
Service: Modulo de Liquidaciones - INVERMIN PAITITI

Formulas basadas en el Excel PL_paititi.xlsx (sheets 'Analisis comercial'):

  TMS          = round(TMH - TMH * %H2O / 100, 3)
  oz_comercial = calcular_ley_comercial(ley_planta, params)  [ya existe en laboratorio.py]
  oz_promedio  = round((oz_comercial + oz_minero) / 2, 4)
  step         = floor(oz_promedio * 10) * 10               [ROUNDDOWN(oz,1)*100 en Excel]
  maquila      = max(95, maquila_base + step)               [col S del Excel]
  insumos      = gasto_acopio + gasto_consumo
  precio_x_tms = ((oz_promedio * rec_liq/100 * (spot - riesgo))
                   - maquila - insumos + bono) * 1.1023     [col AB del Excel]
  total_usd    = max(0, precio_x_tms * tms)                 [col AM del Excel]
  fino_recup   = 31.1035 * 1.1023 * tms * rec_liq/100 * oz_promedio / 100  [col AJ]

REGLA VOLADO (de notas de liquidacion):
  Ley Oz/tc < 0.100 => lote volado, habilitacion a ruma a los 30 dias
"""

from __future__ import annotations

import io
import os
import tempfile
from datetime import date, datetime
from decimal import ROUND_DOWN, ROUND_HALF_UP, Decimal
from typing import Any

import msoffcrypto

# pyrefly: ignore [missing-import]
from app.models.enums import EstadoLiquidacion, EstadoLote, TipoAnalisis
from app.models.models import (
    AnalisisLey,
    AnalisisRecuperacion,
    Liquidacion,
    LiquidacionLote,
    Lote,
    Muestreo,
    ParametrosComerciales,
    Pesaje,
    ProveedorAcopiador,
    PruebaMetalurgica,
    SesionDescarga,
)
from app.schemas.liquidaciones import (
    AlertaLote,
    LiquidacionCreate,
    LiquidacionDetalleOut,
    LiquidacionLoteOut,
    LiquidacionLoteParamsUpdate,
    LiquidacionPreviewOut,
    LiquidacionPreviewRequest,
    LiquidacionResumenOut,
    LoteFinancieroOut,
)
from app.services.config_calculo import get_constantes
from app.services.laboratorio import calcular_ley_comercial, obtener_ley_ag_vigente
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

FACTOR = Decimal("1.1023")
TROY_OZ = Decimal("31.1035")
MIN_MAQUILA = Decimal("95")


def _nombre_entidad(entidad) -> str:
    return entidad.razon_social if entidad else "—"


# ── Helpers de calculo (formulas del Excel) ────────────────────────────────────


def _calc_maquila(oz_promedio: Decimal, maquila_base: Decimal) -> Decimal:
    """
    Excel: =SI((maquila_base + REDONDEAR.MENOS(oz,1)*100) < 95; 95; ...)
    REDONDEAR.MENOS(oz, 1) = truncar a 1 decimal → * 100
    """
    step = oz_promedio.quantize(Decimal("0.1"), rounding=ROUND_DOWN) * 100
    return max(MIN_MAQUILA, maquila_base + step)


def _calc_precio_x_tms(
    oz_promedio: Decimal,
    rec_liq: Decimal,
    spot: Decimal,
    riesgo: Decimal,
    maquila: Decimal,
    insumos: Decimal,
    bono: Decimal,
) -> Decimal:
    """col AB del Excel: ((oz*rec/100*(spot-riesgo)) - maquila - insumos + bono) * factor"""
    val_1 = oz_promedio * rec_liq / 100 * (spot - riesgo)
    val = val_1 - maquila - insumos + bono
    return (val * FACTOR).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _calc_total(precio_x_tms: Decimal, tms: Decimal) -> Decimal:
    """col AM del Excel: max(0, precio_x_tms * tms)"""
    return max(Decimal("0"), (precio_x_tms * tms).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _calc_fino_recuperable(tms: Decimal, rec_liq: Decimal, oz_promedio: Decimal) -> Decimal:
    """col AJ del Excel: 31.1035 * 1.1023 * tms * rec_liq/100 * oz_promedio / 100"""
    return (TROY_OZ * FACTOR * tms * rec_liq / 100 * oz_promedio / 100).quantize(
        Decimal("0.0001"), rounding=ROUND_HALF_UP
    )


# ── Helpers de datos ───────────────────────────────────────────────────────────


def _datos_muestreo_promedio(db: Session, lote_id: int) -> tuple[Decimal, Decimal] | None:
    """
    Retorna (humedad_promedio, tms_promedio) calculados en base a todos los ensayos válidos.
    Se requiere el peso_neto (TMH) de la balanza para calcular el TMS.
    """
    muestreos = db.query(Muestreo).filter(Muestreo.lote_id == lote_id).all()
    if not muestreos:
        return None

    pesaje = db.query(Pesaje).filter(Pesaje.lote_id == lote_id).first()
    if not pesaje or not pesaje.peso_neto:
        return None

    total_h2o = 0.0
    valid_count = 0
    for m in muestreos:
        ph = float(m.peso_humedo) if m.peso_humedo else 0.0
        ps = float(m.peso_seco) if m.peso_seco else 0.0
        if ph > 0 and ps > 0 and ps < ph:
            total_h2o += ((ph - ps) / ph) * 100
            valid_count += 1

    if valid_count > 0:
        h2o_porc = round(total_h2o / valid_count, 2)
        humedad_dec = Decimal(str(h2o_porc))

        tmh = float(pesaje.peso_neto)
        tms_val = round(tmh * (1 - (h2o_porc / 100)), 3)
        tms_dec = Decimal(str(tms_val))
        return (humedad_dec, tms_dec)

    return None


def _pesaje_principal(db: Session, lote_id: int) -> Pesaje | None:
    return db.query(Pesaje).filter(Pesaje.lote_id == lote_id).order_by(Pesaje.id).first()


def _ley_minero(db: Session, lote_id: int) -> Decimal | None:
    a = (
        db.query(AnalisisLey)
        .filter(
            AnalisisLey.lote_id == lote_id,
            AnalisisLey.tipo_analisis == TipoAnalisis.MINERO,
            AnalisisLey.material == "Au",
            AnalisisLey.vigente == True,  # noqa: E712
        )
        .order_by(AnalisisLey.id.desc())
        .first()
    )
    if not a or a.ley_final is None:
        return None
    from app.services.config_calculo import get_constantes, get_quantize_decimal

    constantes = get_constantes(db)
    q_lab = get_quantize_decimal(constantes.decimales_ley_laboratorio)
    return Decimal(str(a.ley_final)).quantize(q_lab, rounding=constantes.redondeo_ley_laboratorio)


def _ley_base_planta(db: Session, lote_id: int) -> Decimal | None:
    """Promedio de análisis planta/externo vigentes. No incluye dirimencia."""
    from app.services.pruebas import calcular_ley_planta

    return calcular_ley_planta(db, lote_id)


def _ley_dirimencia_raw(db: Session, lote_id: int) -> Decimal | None:
    """Ley final del análisis de dirimencia vigente, si existe."""
    a = (
        db.query(AnalisisLey)
        .filter(
            AnalisisLey.lote_id == lote_id,
            AnalisisLey.tipo_analisis == TipoAnalisis.DIRIMENCIA,
            AnalisisLey.material == "Au",
            AnalisisLey.vigente == True,  # noqa: E712
        )
        .first()
    )
    if not a or a.ley_final is None:
        return None
    from app.services.config_calculo import get_constantes, get_quantize_decimal

    constantes = get_constantes(db)
    q_lab = get_quantize_decimal(constantes.decimales_ley_laboratorio)
    return Decimal(str(a.ley_final)).quantize(q_lab, rounding=constantes.redondeo_ley_laboratorio)


def _determinar_rec_liq(
    oz_promedio: Decimal,
    params: ParametrosComerciales | None,
    db: Session,
    lote_id: int,
) -> Decimal | None:
    """
    % Recuperación para LIQUIDAR (col '% Rec Liq' del Excel).
    Valor escalonado según umbrales de ley del acopiador:
      oz_promedio >= umbral_recup_medio → recuperación = 90%
      oz_promedio < umbral_recup_medio → recuperación = 85%
      oz_promedio < umbral_recup_bajo → recuperación = 80%
    El valor 80 para ley muy baja se logra poniendo umbral_recup_bajo=80 en parámetros.

    Fallback: si faltan parámetros/umbrales → usa recuperación real del lab.
    """
    if params and params.umbral_recup_bajo is not None and params.umbral_recup_medio is not None:
        bajo = 80
        medio = 85
        alto = 90

        if oz_promedio >= params.umbral_recup_medio:
            return alto
        elif oz_promedio >= params.umbral_recup_bajo:
            return medio
        elif oz_promedio < params.umbral_recup_bajo:
            return bajo

    # Fallback: recuperación real del laboratorio
    a = (
        db.query(AnalisisRecuperacion)
        .filter(
            AnalisisRecuperacion.lote_id == lote_id,
            AnalisisRecuperacion.vigente == True,  # noqa: E712
            AnalisisRecuperacion.estado != "CERT_RECONOCIMIENTO",
        )
        .order_by(AnalisisRecuperacion.id.desc())
        .first()
    )
    return Decimal(str(a.recuperacion)) if a and a.recuperacion else None


def _rec_planta(db: Session, lote_id: int) -> Decimal | None:
    """% Recuperación REAL del laboratorio (col '% Rec Planta' del Excel)."""
    a = (
        db.query(AnalisisRecuperacion)
        .filter(
            AnalisisRecuperacion.lote_id == lote_id,
            AnalisisRecuperacion.vigente == True,  # noqa: E712
            AnalisisRecuperacion.estado != "CERT_RECONOCIMIENTO",
        )
        .order_by(AnalisisRecuperacion.id.desc())
        .first()
    )
    return Decimal(str(a.recuperacion)) if a and a.recuperacion else None


def _fecha_recepcion(lote: Lote) -> date | None:
    if lote.pesajes:
        dt = lote.pesajes[0].fecha_fin
        return dt.date() if isinstance(dt, datetime) else dt
    return None


def _numero_liquidacion(db: Session) -> str:
    """Genera numero correlativo LIQ-YYYY-NNNN."""
    anio = datetime.now().year
    prefix = f"LIQ-{anio}-"
    ultimo = (
        db.query(Liquidacion.numero_liquidacion)
        .filter(Liquidacion.numero_liquidacion.like(f"{prefix}%"))
        .order_by(Liquidacion.id.desc())
        .first()
    )
    if ultimo and ultimo[0]:
        try:
            n = int(ultimo[0].split("-")[-1]) + 1
        except (ValueError, IndexError):
            n = 1
    else:
        n = 1
    return f"{prefix}{n:04d}"


# ── Calculo de snapshot de un lote ────────────────────────────────────────────


def _calcular_lote(
    db: Session,
    lote: Lote,
    spot_usd: Decimal,
    bono: Decimal,
    rec_liq_override: Decimal | None,
    gasto_acopio_override: Decimal | None = None,
    gasto_consumo_override: Decimal | None = None,
    spot_ag_usd: Decimal | None = None,
) -> tuple[dict[str, Any], list[AlertaLote]]:
    """
    Calcula todos los valores financieros para un lote.
    Retorna (snapshot_dict, alertas).
    alertas con critico=True bloquean la liquidacion.
    """
    alertas: list[AlertaLote] = []
    params: ParametrosComerciales | None = lote.sesion.provacop.parametros

    # ── TMS / TMH ─────────────────────────────────────────────────────────────
    datos_muestreo = _datos_muestreo_promedio(db, lote.id)
    pesaje = _pesaje_principal(db, lote.id)

    if not datos_muestreo:
        alertas.append(
            AlertaLote(
                tipo="SIN_MUESTREO",
                mensaje=f"{lote.ip}: sin muestreo registrado o datos incompletos",
                critico=True,
            )
        )
        return {}, alertas

    humedad, tms = datos_muestreo
    tmh = (
        (tms / (1 - humedad / 100)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        if humedad < 100
        else tms
    )
    sacos = pesaje.sacos if pesaje else None
    fecha_rec = _fecha_recepcion(lote)

    # ── Parametros comerciales ────────────────────────────────────────────────
    if not params:
        alertas.append(
            AlertaLote(
                tipo="SIN_PARAMS", mensaje=f"{lote.ip}: sin parametros comerciales", critico=True
            )
        )
        return {}, alertas

    riesgo = Decimal(str(params.riesgo_comercial)) if params.riesgo_comercial else Decimal("0")
    maquila_base = Decimal(str(params.maquila)) if params.maquila else Decimal("0")
    is_llampo = lote.tipo_material and lote.tipo_material.strip().lower() == "llampo"

    gasto_acopio_base = (
        params.gasto_acopio_llampo
        if is_llampo and params.gasto_acopio_llampo is not None
        else params.gasto_acopio
    )
    gasto_consumo_base = (
        params.gasto_consumo_llampo
        if is_llampo and params.gasto_consumo_llampo is not None
        else params.gasto_consumo
    )

    gasto_acopio = (
        Decimal(str(gasto_acopio_override))
        if gasto_acopio_override is not None
        else (Decimal(str(gasto_acopio_base)) if gasto_acopio_base else Decimal("0"))
    )
    gasto_consumo = (
        Decimal(str(gasto_consumo_override))
        if gasto_consumo_override is not None
        else (Decimal(str(gasto_consumo_base)) if gasto_consumo_base else Decimal("0"))
    )
    insumos = gasto_acopio + gasto_consumo

    # ── Ley Base Planta (solo planta/externo) y dirimencia ───────────────────
    ley_paititi_raw = _ley_base_planta(db, lote.id)
    ley_dirim_raw = _ley_dirimencia_raw(db, lote.id)

    if ley_paititi_raw is None and ley_dirim_raw is None:
        alertas.append(
            AlertaLote(
                tipo="SIN_LEY_PLANTA",
                mensaje=f"{lote.ip}: sin analisis de ley vigente",
                critico=True,
            )
        )
        return {}, alertas

    # ── Configuraciones y Rounding ──────────────────────────────────────────
    constantes = get_constantes(db)
    from app.services.config_calculo import get_quantize_decimal

    q_planta = get_quantize_decimal(constantes.decimales_ley_planta)
    q_comercial = get_quantize_decimal(constantes.decimales_ley_comercial)
    q_final = get_quantize_decimal(constantes.decimales_ley_final)

    # ley_planta para snapshot: ley paititi base (o dirimencia como fallback)
    ley_planta = ley_paititi_raw if ley_paititi_raw is not None else ley_dirim_raw
    if ley_planta is not None:
        ley_planta = Decimal(str(ley_planta)).quantize(
            q_planta, rounding=constantes.redondeo_ley_planta
        )

    # ── Ley Comercial (parámetros aplicados a ley paititi) ───────────────────
    lc_result = calcular_ley_comercial(
        ley_planta,
        params,
        q_comercial=q_comercial,
        rounding=constantes.redondeo_ley_comercial,
    )
    oz_tc_comercial = max(
        Decimal("0"),
        Decimal(str(lc_result["ley_comercial"])).quantize(
            q_comercial, rounding=constantes.redondeo_ley_comercial
        ),
    )

    # ── Auto-set volado + regla: volado → ley comercial = 0 ──────────────────
    if not lote.volado and oz_tc_comercial < constantes.umbral_volado_oz_tc:
        lote.volado = True  # se persiste al commit de crear_liquidacion
    if lote.volado:
        oz_tc_comercial = Decimal("0")

    # ── Ley Minero ────────────────────────────────────────────────────────────
    oz_tc_minero = _ley_minero(db, lote.id)
    if oz_tc_minero is None:
        oz_tc_minero = Decimal("0.00")
        alertas.append(
            AlertaLote(
                tipo="SIN_LEY_MINERO",
                mensaje=f"{lote.ip}: sin ley del minero registrada",
                critico=False,
            )
        )

    # ── Promedio con lógica de dirimencia ────────────────────────────────────
    # Sin dirimencia: promedio simple (paititi + minero) / 2
    # Con dirimencia: clamp(dirimencia, min(paititi,minero), max(paititi,minero))
    #   - dirimencia entre ambos  → se usa dirimencia directamente
    #   - dirimencia > ley_minero → se usa ley_minero (cap)
    #   - dirimencia < ley_paititi → se usa ley_paititi (piso)
    if ley_dirim_raw is not None and oz_tc_minero:
        lc_dir = calcular_ley_comercial(
            ley_dirim_raw,
            params,
            q_comercial=q_comercial,
            rounding=constantes.redondeo_ley_comercial,
        )
        oz_tc_dirimencia = max(
            Decimal("0"),
            Decimal(str(lc_dir["ley_comercial"])).quantize(
                q_comercial, rounding=constantes.redondeo_ley_comercial
            ),
        )
        ley_low = min(oz_tc_comercial, oz_tc_minero)
        ley_high = max(oz_tc_comercial, oz_tc_minero)
        oz_promedio = max(ley_low, min(ley_high, oz_tc_dirimencia)).quantize(
            q_final, rounding=constantes.redondeo_ley_final
        )
    elif oz_tc_minero:
        oz_promedio = ((oz_tc_comercial + oz_tc_minero) / 2).quantize(
            q_final, rounding=constantes.redondeo_ley_final
        )
    else:
        oz_promedio = Decimal(str(oz_tc_comercial)).quantize(
            q_final, rounding=constantes.redondeo_ley_final
        )

    # ── Recuperacion ──────────────────────────────────────────────────────────
    if rec_liq_override is not None:
        rec_liq = rec_liq_override
    else:
        rec_liq = _determinar_rec_liq(oz_promedio, params, db, lote.id)

    if rec_liq is None:
        alertas.append(
            AlertaLote(
                tipo="SIN_RECUPERACION",
                mensaje=f"{lote.ip}: sin analisis de recuperacion vigente",
                critico=True,
            )
        )
        return {}, alertas
    rec_planta_val = _rec_planta(db, lote.id) or rec_liq

    # ── Maquila y precio ──────────────────────────────────────────────────────
    maquila = _calc_maquila(oz_promedio, maquila_base)
    precio_x_tms = _calc_precio_x_tms(
        oz_promedio, rec_liq, spot_usd, riesgo, maquila, insumos, bono
    )
    total_usd = _calc_total(precio_x_tms, tms)
    fino_recuperable = _calc_fino_recuperable(tms, rec_liq, oz_promedio)

    # ── Plata (Ag) ────────────────────────────────────────────────────────────
    # Parámetros contractuales (todos deben estar presentes para pagar Ag)
    ag_result = obtener_ley_ag_vigente(db, lote.id)
    ley_ag_gr_tm: Decimal | None = None
    ley_ag_oz_tc: Decimal | None = None
    valor_ag_usd = Decimal("0")
    aplica_ag = False

    umbral_ag = Decimal(str(params.umbral_ag_oz_tc)) if params and params.umbral_ag_oz_tc else None
    rec_ag = Decimal(str(params.rec_ag_pct)) if params and params.rec_ag_pct else None
    dscto_ag = (
        Decimal(str(params.descuento_ag_usd))
        if params and params.descuento_ag_usd is not None
        else None
    )

    params_ag_completos = all(v is not None for v in (umbral_ag, rec_ag, dscto_ag))

    if ag_result is not None and params_ag_completos:
        ley_ag_gr_tm, ley_ag_oz_tc = ag_result
        if ley_ag_oz_tc >= umbral_ag:
            aplica_ag = True
            if spot_ag_usd:
                # oz_ag_pagables = TMS × ley_ag_oz_tc × (rec_ag / 100)
                oz_ag_pagables = (tms * ley_ag_oz_tc * rec_ag / 100).quantize(
                    Decimal("0.0001"), rounding=ROUND_HALF_UP
                )
                precio_neto_ag = spot_ag_usd - dscto_ag
                if precio_neto_ag > 0:
                    valor_ag_usd = (oz_ag_pagables * precio_neto_ag).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )
        # ley_ag_gr_tm se expone siempre que exista análisis (para mostrar en UI aunque no se pague)

    # ── Profits Operativos ────────────────────────────────────────────────────
    constantes = get_constantes(db)
    costo_fijo_planta = constantes.costo_fijo_planta_maquila or Decimal("80")

    # Prueba Metalurgica para insumos (Soda y Cianuro)
    pm = (
        db.query(PruebaMetalurgica)
        .filter(PruebaMetalurgica.lote_id == lote.id, ~PruebaMetalurgica.descartado)
        .order_by(PruebaMetalurgica.id.desc())
        .first()
    )
    soda_w = Decimal(str(pm.adicion_naoh)) if pm and pm.adicion_naoh else Decimal("3.0")
    cianuro_x = Decimal(str(pm.adicion_nacn)) if pm and pm.adicion_nacn else Decimal("3.0")

    # Evaluacion de proveedor para formula de consumo
    proveedor_nombre = (
        lote.sesion.provacop.acopiador.razon_social if lote.sesion.provacop.acopiador else ""
    ).upper()
    if not proveedor_nombre and lote.sesion.provacop.proveedor:
        proveedor_nombre = lote.sesion.provacop.proveedor.razon_social.upper()

    es_untuca = "UNTUCA" in proveedor_nombre

    cond_untuca = ((soda_w - 3) * Decimal("2.5") + (cianuro_x - 3) * Decimal("3.1")) > 0
    cond_normal = ((soda_w - 3) * Decimal("1.5") + (cianuro_x - 3) * Decimal("3.1")) > 0

    if es_untuca:
        descuento_consumo = (
            ((soda_w - 3) * Decimal("1.5") + (cianuro_x - 3) * Decimal("3.1"))
            if cond_untuca
            else Decimal("0")
        )
    else:
        descuento_consumo = (
            ((soda_w - 3) * Decimal("1.5") + (cianuro_x - 3) * Decimal("3.1"))
            if cond_normal
            else Decimal("0")
        )

    profit_maquila = (maquila - bono) * FACTOR - costo_fijo_planta
    profit_rec = (rec_planta_val - rec_liq) * (spot_usd - riesgo) * ley_planta * FACTOR / 100
    profit_consumo = (gasto_acopio * FACTOR - Decimal("70")) + FACTOR * (
        gasto_consumo - descuento_consumo
    )
    profit_leyes = (ley_planta - oz_promedio) * (spot_usd - riesgo) * rec_planta_val * FACTOR / 100
    profit_total = profit_maquila + profit_rec + profit_consumo + profit_leyes

    # ── Alertas no criticas ───────────────────────────────────────────────────
    if lote.volado:
        dias = (date.today() - fecha_rec).days if fecha_rec else 0
        alertas.append(
            AlertaLote(
                tipo="VOLADO",
                mensaje=f"{lote.ip}: lote VOLADO (ley < 0.100 Oz/tc). Dias en almacen: {dias}",
                critico=False,
            )
        )
    if fecha_rec and (date.today() - fecha_rec).days >= 25:
        dias = (date.today() - fecha_rec).days
        alertas.append(
            AlertaLote(
                tipo="VENCIMIENTO_30D",
                mensaje=f"{lote.ip}: {dias} dias en almacen (limite 30 dias)",
                critico=False,
            )
        )

    snapshot = {
        "ip": lote.ip,
        "fecha_recepcion": fecha_rec,
        "tmh": tmh,
        "pct_humedad": humedad,
        "tms": tms,
        "sacos": sacos,
        "oz_tc_planta": ley_planta,
        "oz_tc_planta_raw": ley_planta,
        "oz_tc_comercial": oz_tc_comercial,
        "oz_tc_minero": oz_tc_minero,
        "oz_tc_promedio": oz_promedio,
        "pct_rec_liq": rec_liq,
        "pct_rec_planta": rec_planta_val,
        "maquila": maquila,
        "riesgo": riesgo,
        "spot_usd": spot_usd,
        "insumos_acopio": gasto_acopio,
        "insumos_consumo": gasto_consumo,
        "insumos_total": insumos,
        "bono": bono,
        "factor": FACTOR,
        "precio_x_tms": precio_x_tms,
        "total_usd": total_usd,
        "fino_recuperable": fino_recuperable,
        "usa_dirimencia": lote.dirimencia,
        "alertas": alertas,
        # Ag
        "ley_ag_gr_tm": ley_ag_gr_tm if aplica_ag else None,
        "ley_ag_oz_tc": ley_ag_oz_tc if aplica_ag else None,
        "spot_ag_usd": spot_ag_usd if aplica_ag else None,
        "valor_ag_usd": valor_ag_usd if aplica_ag else None,
        "aplica_ag": aplica_ag,
        # Profits
        "profit_maquila": profit_maquila,
        "profit_rec": profit_rec,
        "profit_consumo": profit_consumo,
        "profit_leyes": profit_leyes,
        "profit_total": profit_total,
        # campos para el modelo LiquidacionLote
        "tms_snapshot": tms,
        "tmh_snapshot": tmh,
        "humedad_snapshot": humedad,
        "sacos_snapshot": sacos,
        "fecha_recepcion_lote": fecha_rec,
        "maquila_aplicada": maquila,
        "riesgo_aplicado": riesgo,
        "spot_usd_snapshot": spot_usd,
        "insumos_liquidacion": insumos,
        "gasto_acopio_liquidacion": gasto_acopio,
    }

    return snapshot, alertas


# ── API publica del servicio ───────────────────────────────────────────────────


def preview_liquidacion(
    db: Session,
    req: LiquidacionPreviewRequest,
) -> LiquidacionPreviewOut:
    """
    Calcula preview sin guardar. Usado antes de confirmar la liquidacion.
    """
    # 1. Consulta del ProveedorAcopiador usando tus relaciones exactas
    provacop: ProveedorAcopiador = (
        db.query(ProveedorAcopiador)
        .options(
            joinedload(ProveedorAcopiador.proveedor),
            joinedload(ProveedorAcopiador.acopiador),
            joinedload(ProveedorAcopiador.parametros),  # Basado en la línea 192 de tu models.py
        )
        .filter(ProveedorAcopiador.id == req.provacop_id)
        .first()
    )

    if not provacop:
        raise ValueError(f"ProveedorAcopiador {req.provacop_id} no encontrado")

    spot = req.spot_usd
    lotes_out: list[LoteFinancieroOut] = []
    alertas_globales: list[AlertaLote] = []
    puede_generar = True

    for item in req.lotes:
        # 2. Consulta del Lote encadenando las relaciones hasta llegar a parámetros
        lote = (
            db.query(Lote)
            .options(
                joinedload(Lote.pesajes),
                joinedload(Lote.muestreos),
                joinedload(Lote.sesion)
                .joinedload(SesionDescarga.provacop)
                .joinedload(ProveedorAcopiador.parametros),
            )
            .filter(Lote.ip == item.ip, ~Lote.eliminado)
            .first()
        )

        if not lote:
            alertas_globales.append(
                AlertaLote(
                    tipo="LOTE_NO_ENCONTRADO", mensaje=f"Lote {item.ip} no encontrado", critico=True
                )
            )
            puede_generar = False
            continue

        snap, alertas = _calcular_lote(
            db,
            lote,
            spot,
            item.bono or Decimal("0"),
            item.rec_liq_override,
            item.gasto_acopio_override,
            item.gasto_consumo_override,
            spot_ag_usd=req.spot_ag_usd,
        )

        if alertas:
            for alerta in alertas:
                # Le añadimos el IP al inicio del mensaje para saber de qué lote proviene el error
                alerta.mensaje = f"[{item.ip}] {alerta.mensaje}"
                alertas_globales.append(alerta)

        if any(a.critico for a in alertas):
            puede_generar = False

        if not snap:
            continue

        lotes_out.append(
            LoteFinancieroOut(
                **{k: v for k, v in snap.items() if k in LoteFinancieroOut.model_fields}
            )
        )

    total_usd = sum(lo.total_usd for lo in lotes_out)
    total_tms = sum(lo.tms for lo in lotes_out)
    total_tmh = sum(lo.tmh for lo in lotes_out)
    total_oz = sum(lo.fino_recuperable for lo in lotes_out)
    total_ag = sum((lo.valor_ag_usd or Decimal("0")) for lo in lotes_out)

    return LiquidacionPreviewOut(
        provacop_id=req.provacop_id,
        proveedor_razon_social=provacop.proveedor.razon_social if provacop.proveedor else "-",
        proveedor_ruc=provacop.proveedor.ruc if provacop.proveedor else None,
        acopiador_nombre=provacop.acopiador.razon_social if provacop.acopiador else "-",
        spot_usd=spot,
        lotes=lotes_out,
        total_usd=total_usd,
        total_tms=total_tms,
        total_tmh=total_tmh,
        total_oz_compradas=total_oz,
        count_lotes=len(lotes_out),
        alertas_globales=alertas_globales,
        puede_generar=puede_generar and bool(lotes_out),
        total_ag_usd=total_ag,
        hay_ag=any(lo.aplica_ag for lo in lotes_out),
    )


def crear_liquidacion(
    db: Session,
    req: LiquidacionCreate,
    usuario_id: int,
) -> Liquidacion:
    """
    Crea la liquidacion con snapshot completo de valores financieros.
    Actualiza estado de lotes a LIQUIDADO.
    """
    provacop = db.query(ProveedorAcopiador).filter(ProveedorAcopiador.id == req.provacop_id).first()
    if not provacop:
        raise ValueError(f"ProveedorAcopiador {req.provacop_id} no encontrado")

    numero = req.numero_liquidacion or _numero_liquidacion(db)

    liq = Liquidacion(
        numero_liquidacion=numero,
        provacop_id=req.provacop_id,
        precio_oro_usd=req.spot_usd,
        estado=EstadoLiquidacion.BORRADOR if req.como_borrador else EstadoLiquidacion.GENERADA,
        creado_por=usuario_id,
    )
    db.add(liq)
    db.flush()

    total_general = Decimal("0")

    for item in req.lotes:
        lote = (
            db.query(Lote)
            .options(
                joinedload(Lote.pesajes),
                joinedload(Lote.muestreos),
                joinedload(Lote.sesion)
                .joinedload(SesionDescarga.provacop)
                .joinedload(ProveedorAcopiador.parametros),
            )
            .filter(Lote.ip == item.ip, Lote.eliminado == False)  # noqa: E712
            .first()
        )
        if not lote:
            raise ValueError(f"Lote {item.ip} no encontrado")

        snap, alertas = _calcular_lote(
            db,
            lote,
            req.spot_usd,
            item.bono or Decimal("0"),
            item.rec_liq_override,
            item.gasto_acopio_override,
            item.gasto_consumo_override,
            spot_ag_usd=req.spot_ag_usd,
        )

        if any(a.critico for a in alertas):
            msgs = [a.mensaje for a in alertas if a.critico]
            raise ValueError(f"No se puede liquidar {item.ip}: {'; '.join(msgs)}")

        ll = LiquidacionLote(
            liquidacion_id=liq.id,
            lote_id=lote.id,
            fecha_emision=req.fecha_liquidacion or date.today(),
            fecha_recepcion=snap["fecha_recepcion"],
            fecha_recepcion_lote=snap["fecha_recepcion"],
            ley_comercial=snap["oz_tc_promedio"],
            usa_dirimencia=snap["usa_dirimencia"],
            oz_tc_planta=snap["oz_tc_planta"],
            oz_tc_comercial=snap["oz_tc_comercial"],
            oz_tc_minero=snap["oz_tc_minero"],
            oz_tc_promedio=snap["oz_tc_promedio"],
            porcentaje_rec_liquido=snap["pct_rec_liq"],
            porcentaje_rec_planta=snap["pct_rec_planta"],
            fino_recuperable=snap["fino_recuperable"],
            gasto_acopio_liquidacion=snap["insumos_acopio"],
            bono=snap["bono"],
            insumos_liquidacion=snap["insumos_total"],
            maquila_aplicada=snap["maquila"],
            riesgo_aplicado=snap["riesgo"],
            spot_usd_snapshot=snap["spot_usd"],
            precio_x_tms=snap["precio_x_tms"],
            total_usd=snap["total_usd"],
            tms_snapshot=snap["tms"],
            tmh_snapshot=snap["tmh"],
            humedad_snapshot=snap["pct_humedad"],
            sacos_snapshot=snap["sacos"],
            ley_ag_gr_tm_snapshot=snap.get("ley_ag_gr_tm"),
            spot_ag_snapshot=snap.get("spot_ag_usd"),
            valor_ag_usd=snap.get("valor_ag_usd"),
            creado_por=usuario_id,
        )
        db.add(ll)

        # Solo actualizar estado del lote si NO es borrador
        if not req.como_borrador:
            lote.estado = EstadoLote.LIQUIDADO
            lote.estado_modificado_por = usuario_id
            lote.fecha_modificacion_estado = datetime.now()

        total_general += snap["total_usd"]

    liq.valor_total_usd = total_general
    db.commit()
    db.refresh(liq)
    return liq


def obtener_liquidaciones(
    db: Session,
    provacop_id: int | None = None,
    estado: str | None = None,
) -> list[LiquidacionResumenOut]:
    q = (
        db.query(Liquidacion)
        .options(
            joinedload(Liquidacion.provacop).joinedload(ProveedorAcopiador.proveedor),
            joinedload(Liquidacion.provacop).joinedload(ProveedorAcopiador.acopiador),
            joinedload(Liquidacion.liquidacion_lotes),
        )
        .order_by(Liquidacion.id.desc())
    )
    if provacop_id:
        q = q.filter(Liquidacion.provacop_id == provacop_id)
    if estado:
        q = q.filter(Liquidacion.estado == estado)

    result = []
    for liq in q.all():
        result.append(_to_resumen(liq))
    return result


def obtener_liquidacion(db: Session, liquidacion_id: int) -> LiquidacionDetalleOut | None:
    liq = (
        db.query(Liquidacion)
        .options(
            joinedload(Liquidacion.provacop).joinedload(ProveedorAcopiador.proveedor),
            joinedload(Liquidacion.provacop).joinedload(ProveedorAcopiador.acopiador),
            joinedload(Liquidacion.liquidacion_lotes)
            .joinedload(LiquidacionLote.lote)
            .joinedload(Lote.pesajes),
        )
        .filter(Liquidacion.id == liquidacion_id)
        .first()
    )
    if not liq:
        return None
    return _to_detalle(liq)


def cambiar_estado(
    db: Session,
    liquidacion_id: int,
    nuevo_estado: str,
    usuario_id: int,
) -> Liquidacion:
    liq = db.query(Liquidacion).filter(Liquidacion.id == liquidacion_id).first()
    if not liq:
        raise ValueError("Liquidacion no encontrada")
    if liq.estado == EstadoLiquidacion.PAGADA:
        raise ValueError("Una liquidacion PAGADA no puede modificarse")

    liq.estado = nuevo_estado
    if nuevo_estado == EstadoLiquidacion.PAGADA:
        liq.cerrado_por = usuario_id
        liq.fecha_cierre = datetime.now()

    # Actualizar estado de lotes
    for ll in liq.liquidacion_lotes:
        lote = ll.lote
        if lote:
            if nuevo_estado == EstadoLiquidacion.FACTURADA:
                lote.estado = EstadoLote.FACTURADO
                lote.estado_modificado_por = usuario_id
                lote.fecha_modificacion_estado = datetime.now()
            elif nuevo_estado == EstadoLiquidacion.PAGADA:
                lote.estado = EstadoLote.PAGADO
                lote.estado_modificado_por = usuario_id
                lote.fecha_modificacion_estado = datetime.now()

    db.commit()
    db.refresh(liq)
    return liq


def emitir_liquidacion(
    db: Session,
    liquidacion_id: int,
    usuario_id: int,
) -> Liquidacion:
    """
    Transicion BORRADOR → GENERADA.
    Actualiza lotes a LIQUIDADO y recalcula totales desde snapshot guardado.
    """
    liq = (
        db.query(Liquidacion)
        .options(joinedload(Liquidacion.liquidacion_lotes).joinedload(LiquidacionLote.lote))
        .filter(Liquidacion.id == liquidacion_id)
        .first()
    )
    if not liq:
        raise ValueError("Liquidacion no encontrada")
    if liq.estado != EstadoLiquidacion.BORRADOR:
        raise ValueError(
            f"Solo se puede emitir una liquidacion en estado BORRADOR, estado actual: {liq.estado}"
        )

    for ll in liq.liquidacion_lotes:
        lote = ll.lote
        if not lote:
            continue
        if lote.estado != EstadoLote.RECEPCIONADO:
            raise ValueError(
                f"Lote {lote.ip} ya no está en estado RECEPCIONADO (estado: {lote.estado})"
            )
        lote.estado = EstadoLote.LIQUIDADO
        lote.estado_modificado_por = usuario_id
        lote.fecha_modificacion_estado = datetime.now()

    liq.estado = EstadoLiquidacion.GENERADA
    db.commit()
    db.refresh(liq)
    return liq


# ── Serializadores internos ───────────────────────────────────────────────────


def _to_resumen(liq: Liquidacion) -> LiquidacionResumenOut:
    prov = liq.provacop.proveedor if liq.provacop else None
    acop = liq.provacop.acopiador if liq.provacop else None
    return LiquidacionResumenOut(
        id=liq.id,
        numero_liquidacion=liq.numero_liquidacion or "",
        estado=liq.estado,
        provacop_id=liq.provacop_id,
        proveedor_razon_social=prov.razon_social if prov else "-",
        proveedor_ruc=prov.ruc if prov else None,
        acopiador_nombre=acop.razon_social if acop else "-",
        spot_usd=liq.precio_oro_usd or Decimal("0"),
        total_usd=liq.valor_total_usd or Decimal("0"),
        count_lotes=len(liq.liquidacion_lotes),
        fecha_creacion=liq.creado_en or datetime.now(),
    )


def _to_lote_out(ll: LiquidacionLote) -> LiquidacionLoteOut:
    return LiquidacionLoteOut(
        liquidacion_id=ll.liquidacion_id,
        ip=ll.lote.ip if ll.lote else "",
        fecha_recepcion=ll.fecha_recepcion_lote,
        fecha_emision=ll.fecha_emision,
        tmh=ll.tmh_snapshot or Decimal("0"),
        pct_humedad=ll.humedad_snapshot or Decimal("0"),
        tms=ll.tms_snapshot or Decimal("0"),
        sacos=ll.sacos_snapshot,
        oz_tc_planta=ll.oz_tc_planta or Decimal("0"),
        oz_tc_comercial=ll.oz_tc_comercial or Decimal("0"),
        oz_tc_minero=ll.oz_tc_minero or Decimal("0"),
        oz_tc_promedio=ll.oz_tc_promedio or Decimal("0"),
        pct_rec_liq=ll.porcentaje_rec_liquido or Decimal("0"),
        pct_rec_planta=ll.porcentaje_rec_planta,
        maquila=ll.maquila_aplicada or Decimal("0"),
        riesgo=ll.riesgo_aplicado or Decimal("0"),
        spot_usd=ll.spot_usd_snapshot or Decimal("0"),
        insumos_acopio=ll.gasto_acopio_liquidacion or Decimal("0"),
        insumos_consumo=(ll.insumos_liquidacion or Decimal("0"))
        - (ll.gasto_acopio_liquidacion or Decimal("0")),
        insumos_total=ll.insumos_liquidacion or Decimal("0"),
        bono=ll.bono or Decimal("0"),
        factor=FACTOR,
        precio_x_tms=ll.precio_x_tms or Decimal("0"),
        total_usd=ll.total_usd or Decimal("0"),
        fino_recuperable=ll.fino_recuperable or Decimal("0"),
        usa_dirimencia=ll.usa_dirimencia or False,
        alertas=[],
        ley_ag_gr_tm=ll.ley_ag_gr_tm_snapshot,
        spot_ag_usd=ll.spot_ag_snapshot,
        valor_ag_usd=ll.valor_ag_usd,
        aplica_ag=bool(ll.valor_ag_usd and ll.valor_ag_usd > 0),
    )


def _to_detalle(liq: Liquidacion) -> LiquidacionDetalleOut:
    resumen = _to_resumen(liq)
    lotes = [_to_lote_out(ll) for ll in liq.liquidacion_lotes]
    return LiquidacionDetalleOut(
        **resumen.model_dump(),
        lotes=lotes,
        pdf_url=liq.pdf_url,
        fecha_cierre=liq.fecha_cierre,
    )


# ── Logica de volado (llamada desde laboratorio al registrar ley) ──────────────


def evaluar_volado(db: Session, lote_id: int, ley_planta: Decimal, usuario_id: int) -> bool:
    """
    Marca el lote como volado si ley_planta < 0.100.
    Retorna True si se marco como volado (nuevo).
    """
    lote = db.query(Lote).filter(Lote.id == lote_id).first()
    if not lote:
        return False
    constantes = get_constantes(db)
    if not lote.volado and ley_planta < constantes.umbral_volado_oz_tc:
        lote.volado = True
        lote.modificado_por = usuario_id
        return True
    return False


def lotes_disponibles_para_liquidar(
    db: Session,
    provacop_id: int,
) -> list[dict]:
    lotes = (
        db.query(Lote)
        .options(joinedload(Lote.pesajes), joinedload(Lote.muestreos))
        .join(Lote.sesion)
        .filter(
            Lote.sesion.has(provacop_id=provacop_id),
            Lote.eliminado == False,  # noqa: E712
            Lote.estado == EstadoLote.RECEPCIONADO,
            Lote.tipo_material.in_(["Mineral", "Llampo", "M.Llampo"]),
        )
        .order_by(Lote.id.desc())
        .all()
    )

    resultado = []
    for lote in lotes:
        ya_liquidado = (
            db.query(LiquidacionLote)
            .join(Liquidacion)
            .filter(
                LiquidacionLote.lote_id == lote.id,
                Liquidacion.estado.notin_([EstadoLiquidacion.BORRADOR]),
            )
            .first()
        )
        if ya_liquidado:
            continue

        fecha_rec = _fecha_recepcion(lote)
        dias = (date.today() - fecha_rec).days if fecha_rec else 0
        datos_muestreo = _datos_muestreo_promedio(db, lote.id)
        pesaje = _pesaje_principal(db, lote.id)

        # ── Ley y recuperación ──────────────────────────────────────
        ley_paititi_r = _ley_base_planta(db, lote.id)
        ley_dirim_r = _ley_dirimencia_raw(db, lote.id)
        oz_tc_planta = float(ley_paititi_r) if ley_paititi_r else None
        oz_tc_minero_val = _ley_minero(db, lote.id)
        oz_tc_minero = float(oz_tc_minero_val) if oz_tc_minero_val else None
        usa_dir = bool(lote.dirimencia)

        constantes = get_constantes(db)
        from app.services.config_calculo import get_quantize_decimal

        q_comercial = get_quantize_decimal(constantes.decimales_ley_comercial)
        q_final = get_quantize_decimal(constantes.decimales_ley_final)
        params = lote.sesion.provacop.parametros if lote.sesion and lote.sesion.provacop else None
        ley_comercial = None
        ley_base = ley_paititi_r if ley_paititi_r is not None else ley_dirim_r

        if ley_base is not None:
            lc = calcular_ley_comercial(
                ley_base,
                params,
                constantes.umbral_volado_oz_tc,
                q_comercial=q_comercial,
                rounding=constantes.redondeo_ley_comercial,
            )
            oz_com_paititi = Decimal(
                str(max(0.0, float(lc["ley_comercial"])) if lc else 0)
            ).quantize(q_comercial, rounding=constantes.redondeo_ley_comercial)

            # Auto-set volado basado en ley comercial paititi
            if not lote.volado and oz_com_paititi < constantes.umbral_volado_oz_tc:
                lote.volado = True
                db.flush()

            if lote.volado:
                ley_comercial = Decimal("0")
            elif ley_dirim_r is not None and oz_tc_minero_val:
                # Dirimencia: clamp(dir, min(paititi,minero), max(paititi,minero))
                lc_dir = calcular_ley_comercial(
                    ley_dirim_r,
                    params,
                    constantes.umbral_volado_oz_tc,
                    q_comercial=q_comercial,
                    rounding=constantes.redondeo_ley_comercial,
                )
                oz_dir = Decimal(
                    str(max(0.0, float(lc_dir["ley_comercial"])) if lc_dir else 0)
                ).quantize(q_comercial, rounding=constantes.redondeo_ley_comercial)
                ley_low = min(oz_com_paititi, oz_tc_minero_val)
                ley_high = max(oz_com_paititi, oz_tc_minero_val)
                ley_comercial = max(ley_low, min(ley_high, oz_dir))
            elif oz_tc_minero_val:
                ley_comercial = ((oz_com_paititi + oz_tc_minero_val) / 2).quantize(
                    q_final, rounding=constantes.redondeo_ley_final
                )
            else:
                ley_comercial = oz_com_paititi

        ley_comercial = (
            Decimal(str(ley_comercial)).quantize(q_final, rounding=constantes.redondeo_ley_final)
            if ley_comercial is not None
            else None
        )
        # rec_liq = _determinar_rec_liq(ley_comercial, params, db, lote.id)
        rec = _rec_planta(db, lote.id)
        porcentaje_rec = float(rec) if rec else None
        listo = all(x is not None for x in [ley_comercial, porcentaje_rec, datos_muestreo])

        resultado.append(
            {
                "ip": lote.ip,
                "tipo_material": lote.tipo_material,
                "fecha_recepcion": fecha_rec,
                "dias_almacen": dias,
                "tms": float(datos_muestreo[1]) if datos_muestreo else None,
                "tmh": float(pesaje.peso_neto) if pesaje and pesaje.peso_neto else None,
                "sacos": pesaje.sacos if pesaje else None,
                "volado": lote.volado,
                "alerta_vencimiento": dias >= 25,
                "ley_comercial": ley_comercial,
                "oz_tc_planta": oz_tc_planta,
                "oz_tc_minero": oz_tc_minero,
                "porcentaje_rec": porcentaje_rec,
                "usa_dirimencia": usa_dir,
                "listo_para_liquidar": listo,
                "provacop_id": provacop_id,  # ya viene como parámetro de la función
                "proveedor": _nombre_entidad(lote.sesion.provacop.proveedor)
                if lote.sesion and lote.sesion.provacop
                else "—",
                "acopiador": _nombre_entidad(lote.sesion.provacop.acopiador)
                if lote.sesion and lote.sesion.provacop
                else "—",
            }
        )

    return resultado


def editar_params_lote(
    db: Session,
    liquidacion_id: int,
    ip: str,
    params: LiquidacionLoteParamsUpdate,
    usuario_id: int,
) -> LiquidacionLote:
    """
    Admin/Gerencia editan parámetros de un lote en una liquidación no PAGADA.
    Recalcula precio_x_tms, total_usd y fino_recuperable.
    """
    liq = db.query(Liquidacion).filter(Liquidacion.id == liquidacion_id).first()
    if not liq:
        raise ValueError("Liquidación no encontrada")
    if liq.estado == EstadoLiquidacion.PAGADA:
        raise ValueError("No se puede modificar una liquidación PAGADA")

    ll = (
        db.query(LiquidacionLote)
        .join(Lote, Lote.id == LiquidacionLote.lote_id)
        .filter(
            LiquidacionLote.liquidacion_id == liquidacion_id,
            Lote.ip == ip,
        )
        .first()
    )
    if not ll:
        raise ValueError(f"Lote {ip} no encontrado en esta liquidación")

    if params.bono is not None:
        ll.bono = params.bono
    if params.rec_liq_override is not None:
        ll.porcentaje_rec_liquido = params.rec_liq_override
    if params.riesgo_override is not None:
        ll.riesgo_aplicado = params.riesgo_override
    if params.maquila_override is not None:
        ll.maquila_aplicada = params.maquila_override
    if params.gasto_acopio_override is not None:
        consumo_prev = (ll.insumos_liquidacion or Decimal("0")) - (
            ll.gasto_acopio_liquidacion or Decimal("0")
        )
        ll.gasto_acopio_liquidacion = params.gasto_acopio_override
        ll.insumos_liquidacion = params.gasto_acopio_override + consumo_prev
    if params.gasto_consumo_override is not None:
        ll.insumos_liquidacion = (
            ll.gasto_acopio_liquidacion or Decimal("0")
        ) + params.gasto_consumo_override

    # Recalcular con valores actualizados
    oz = ll.oz_tc_promedio or Decimal("0")
    rec_liq = ll.porcentaje_rec_liquido or Decimal("0")
    spot = ll.spot_usd_snapshot or Decimal("0")
    riesgo = ll.riesgo_aplicado or Decimal("0")
    maquila = ll.maquila_aplicada or Decimal("0")
    insumos = ll.insumos_liquidacion or Decimal("0")
    bono = ll.bono or Decimal("0")
    tms = ll.tms_snapshot or Decimal("0")

    ll.precio_x_tms = _calc_precio_x_tms(oz, rec_liq, spot, riesgo, maquila, insumos, bono)
    ll.total_usd = _calc_total(ll.precio_x_tms, tms)
    ll.fino_recuperable = _calc_fino_recuperable(tms, rec_liq, oz)
    ll.modificado_por = usuario_id

    db.flush()
    return ll


def generar_excel_pl(db: Session, clave: str) -> io.BytesIO:
    """Genera un Excel con el historial de Lotes liquidados (PL) con contrasena."""
    q = (
        db.query(LiquidacionLote)
        .options(
            joinedload(LiquidacionLote.lote).joinedload(Lote.sesion),
            joinedload(LiquidacionLote.lote).joinedload(Lote.ruma),
            joinedload(LiquidacionLote.liquidacion)
            .joinedload(Liquidacion.provacop)
            .joinedload(ProveedorAcopiador.proveedor),
            joinedload(LiquidacionLote.liquidacion)
            .joinedload(Liquidacion.provacop)
            .joinedload(ProveedorAcopiador.acopiador),
        )
        .order_by(LiquidacionLote.liquidacion_id.desc(), LiquidacionLote.lote_id.desc())
    )
    llotes = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"

    headers = [
        "Lote (IP)",
        "Fecha Emision",
        "Fecha Recep.",
        "Proveedor",
        "RUC",
        "Acopiador",
        "Sacos",
        "Material",
        "TMH (Peso Neto)",
        "% H2O",
        "TMS",
        "Ley Planta (Oz/TC)",
        "Ley Comercial (Oz/TC)",
        "Ley gr/TM",
        "Ley Minero (Oz/TC)",
        "Ley Promedio (Oz/TC)",
        "% Rec Liq",
        "% Rec Planta",
        "Spot USD",
        "Factor",
        "Bono ($/TM)",
        "Riesgo ($/TM)",
        "Maquila ($/TM)",
        "Gasto Acopio ($/TM)",
        "Gasto Consumo ($/TM)",
        "Insumos ($/TM)",
        "Precio x TMS ($/TM)",
        "Bruto Au (USD)",
        "Ley Ag (Gr/TM)",
        "Ley Ag (Oz/TC)",
        "Bruto Ag (USD)",
        "Profit Maquila ($/TM)",
        "Profit Rec ($/TM)",
        "Profit Consumo ($/TM)",
        "Profit Leyes ($/TM)",
        "Profit Total $/TM",
        "Total USD",
        "Total Pago al Minero (USD)",
        "Fino Rec. (Oz)",
        "Ruma",
        "Campana",
        "Estado Lote",
        "N° Liquidacion",
        "Estado Liq.",
        "Placa",
        "Conductor",
        "Transportista",
        "Guia Remision",
        "Guia Transporte",
    ]

    gold_fill = PatternFill("solid", fgColor="1C1A09")
    gold_font = Font(bold=True, color="C9A227")
    center = Alignment(horizontal="center")

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = gold_fill
        c.font = gold_font
        c.alignment = center

    for row_idx, ll in enumerate(llotes, 2):
        liq = ll.liquidacion
        lote = ll.lote
        sesion = lote.sesion if lote else None
        ruma = lote.ruma if lote else None

        prov = (
            str(liq.provacop.proveedor.razon_social)
            if liq.provacop and liq.provacop.proveedor and liq.provacop.proveedor.razon_social
            else ""
        )
        ruc = (
            str(liq.provacop.proveedor.ruc)
            if liq.provacop and liq.provacop.proveedor and liq.provacop.proveedor.ruc
            else ""
        )
        acop = (
            str(liq.provacop.acopiador.razon_social)
            if liq.provacop and liq.provacop.acopiador and liq.provacop.acopiador.razon_social
            else ""
        )

        try:
            calc_dict, _ = _calcular_lote(
                db,
                lote,
                spot_usd=ll.spot_usd_snapshot or Decimal("0"),
                bono=ll.bono or Decimal("0"),
                rec_liq_override=ll.porcentaje_rec_liquido,
                gasto_acopio_override=ll.gasto_acopio_liquidacion,
                gasto_consumo_override=(ll.insumos_liquidacion or Decimal("0"))
                - (ll.gasto_acopio_liquidacion or Decimal("0")),
            )
        except Exception:
            calc_dict = {}

        p_maquila = float(calc_dict.get("profit_maquila") or 0)
        p_rec = float(calc_dict.get("profit_rec") or 0)
        p_consumo = float(calc_dict.get("profit_consumo") or 0)
        p_leyes = float(calc_dict.get("profit_leyes") or 0)
        p_total = float(calc_dict.get("profit_total") or 0)

        oz_tc_planta = float(calc_dict.get("oz_tc_planta") or 0)
        if not oz_tc_planta and lote:
            oz_tc_planta = float(
                next(
                    (
                        a.ley_final
                        for a in lote.analisis_ley
                        if a.vigente and a.tipo_analisis == "planta"
                    ),
                    0,
                )
            )

        oz_tc_comercial = float(ll.oz_tc_comercial or calc_dict.get("oz_tc_comercial") or 0)
        if not oz_tc_comercial and lote:
            oz_tc_comercial = float(
                next(
                    (
                        a.ley_final
                        for a in lote.analisis_ley
                        if a.vigente and a.tipo_analisis == "comercial"
                    ),
                    0,
                )
            )

        ley_gr_tm = float(
            next(
                (
                    a.ley_gr_tm
                    for a in lote.analisis_ley
                    if a.vigente and a.tipo_analisis == "comercial"
                ),
                0,
            )
            if lote
            else 0
        )

        oz_tc_minero = float(calc_dict.get("oz_tc_minero") or 0)
        oz_tc_promedio = float(calc_dict.get("oz_tc_promedio") or 0)
        pct_rec_liq = float(ll.porcentaje_rec_liquido or calc_dict.get("pct_rec_liq") or 0)
        pct_rec_planta = float(calc_dict.get("pct_rec_planta") or 0)
        spot_usd = float(ll.spot_usd_snapshot or calc_dict.get("spot_usd") or 0)
        factor = float(calc_dict.get("factor") or 1.1023)
        bono = float(ll.bono or calc_dict.get("bono") or 0)
        riesgo = float(calc_dict.get("riesgo") or 0)
        maquila = float(ll.maquila_aplicada or calc_dict.get("maquila") or 0)
        gasto_acopio = float(ll.gasto_acopio_liquidacion or calc_dict.get("insumos_acopio") or 0)
        insumos_total = float(ll.insumos_liquidacion or calc_dict.get("insumos_total") or 0)
        gasto_consumo_calc = calc_dict.get("insumos_consumo")
        if gasto_consumo_calc is not None:
            gasto_consumo = float(gasto_consumo_calc)
        else:
            gasto_consumo = float(insumos_total - gasto_acopio)

        precio_x_tms = float(calc_dict.get("precio_x_tms") or 0)
        tms = float(ll.tms_snapshot or 0)
        bruto_au_usd = float(round(precio_x_tms * tms, 2))

        # Plata (Ag)
        ley_ag_gr_tm = float(calc_dict.get("ley_ag_gr_tm") or 0)
        ley_ag_oz_tc = float(calc_dict.get("ley_ag_oz_tc") or 0)
        bruto_ag_usd = float(calc_dict.get("valor_ag_usd") or 0)

        total_usd = float(ll.total_usd or calc_dict.get("total_usd") or 0)
        total_pago_minero = float(round(total_usd + bruto_ag_usd, 2))
        fino_recuperable = float(ll.fino_recuperable or calc_dict.get("fino_recuperable") or 0)

        row_data = [
            lote.ip if lote else "",
            liq.creado_en.strftime("%Y-%m-%d") if liq and liq.creado_en else "",
            ll.fecha_recepcion_lote.strftime("%Y-%m-%d") if ll.fecha_recepcion_lote else "",
            prov,
            ruc,
            acop,
            ll.sacos_snapshot,
            lote.tipo_material if lote else "",
            float(ll.tmh_snapshot or 0),
            float(ll.humedad_snapshot or 0),
            tms,
            oz_tc_planta,
            oz_tc_comercial,
            ley_gr_tm,
            oz_tc_minero,
            oz_tc_promedio,
            pct_rec_liq,
            pct_rec_planta,
            spot_usd,
            factor,
            bono,
            riesgo,
            maquila,
            gasto_acopio,
            gasto_consumo,
            insumos_total,
            precio_x_tms,
            bruto_au_usd,
            ley_ag_gr_tm,
            ley_ag_oz_tc,
            bruto_ag_usd,
            p_maquila,
            p_rec,
            p_consumo,
            p_leyes,
            p_total,
            total_usd,
            total_pago_minero,
            fino_recuperable,
            ruma.codigo if ruma else "",
            ruma.campana if ruma else "",
            lote.estado if lote else "",
            liq.numero_liquidacion or str(liq.id),
            liq.estado,
            sesion.placa if sesion else "",
            sesion.conductor if sesion else "",
            sesion.transportista if sesion else "",
            sesion.guia_remision if sesion else "",
            sesion.guia_transporte if sesion else "",
        ]

        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    out = io.BytesIO()
    with open(tmp_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.encrypt(clave, out)

    os.remove(tmp_path)
    out.seek(0)
    return out
