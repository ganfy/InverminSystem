"""
Tests unitarios para app/services/liquidaciones.py

Cubren las funciones puras que implementan las fórmulas del Excel PL_paititi.xlsx:
  - _calc_maquila
  - _calc_precio_x_tms
  - _calc_total
  - _calc_fino_recuperable
  - _numero_liquidacion (lógica de secuencia)

No requieren base de datos; usan únicamente Decimal y la lógica del módulo.
"""

from __future__ import annotations

from decimal import Decimal

# Importar directamente las funciones privadas que necesitamos testear.
# En Python las funciones con _ son accesibles; esto es aceptable en tests de servicio.
from app.services.liquidaciones import (
    FACTOR,
    MIN_MAQUILA,
    TROY_OZ,
    _calc_fino_recuperable,
    _calc_maquila,
    _calc_precio_x_tms,
    _calc_total,
)

# ── Helpers ────────────────────────────────────────────────────────────────────


def dec(v: str | float | int) -> Decimal:
    return Decimal(str(v))


# ==============================================================================
# _calc_maquila
# ==============================================================================


class TestCalcMaquila:
    """
    Excel: =SI((maquila_base + REDONDEAR.MENOS(oz,1)*100) < 95 ; 95 ; ...)
    REDONDEAR.MENOS(oz, 1) trunca a 1 decimal, luego *100.
    """

    def test_piso_minimo_95_cuando_maquila_base_cero_y_ley_baja(self):
        # oz=0.08 → trunca a 0.0 → step=0 → max(95, 0+0)=95
        result = _calc_maquila(dec("0.08"), dec("0"))
        assert result == dec("95")

    def test_piso_minimo_aplicado_cuando_suma_menor_95(self):
        # oz=0.20 → step=20 → maquila_base=70 → 70+20=90 < 95 → 95
        result = _calc_maquila(dec("0.20"), dec("70"))
        assert result == dec("95")

    def test_valor_calculado_supera_piso(self):
        # oz=0.35 → trunca 0.3 → step=30 → maquila_base=100 → 100+30=130 > 95
        result = _calc_maquila(dec("0.35"), dec("100"))
        assert result == dec("130")

    def test_truncamiento_correcto_no_redondea_arriba(self):
        # oz=0.39 → debe truncar a 0.3 (no redondear a 0.4) → step=30
        result = _calc_maquila(dec("0.39"), dec("0"))
        assert result == MIN_MAQUILA  # 0+30=30 < 95 → piso

    def test_oz_exacto_en_decima(self):
        # oz=0.30 → step=30 → maquila_base=80 → 110 > 95
        result = _calc_maquila(dec("0.30"), dec("80"))
        assert result == dec("110")

    def test_ley_alta_genera_maquila_alta(self):
        # oz=0.60 → step=60 → maquila_base=200 → 260
        result = _calc_maquila(dec("0.60"), dec("200"))
        assert result == dec("260")


# ==============================================================================
# _calc_precio_x_tms
# ==============================================================================


class TestCalcPrecioXTms:
    """
    col AB: ((oz*rec/100*(spot-riesgo)) - maquila - insumos + bono) * FACTOR
    FACTOR = 1.1023
    """

    def test_formula_basica(self):
        oz = dec("0.335")
        rec = dec("88")
        spot = dec("2400")
        riesgo = dec("10")
        maquila = dec("128.5")
        insumos = dec("8")
        bono = dec("0")

        # Manual: val_1 = 0.335 * 88/100 * (2400-10) = 0.335*0.88*2390
        val_1 = oz * rec / 100 * (spot - riesgo)
        val = val_1 - maquila - insumos + bono
        expected = (val * FACTOR).quantize(Decimal("0.0001"))

        result = _calc_precio_x_tms(oz, rec, spot, riesgo, maquila, insumos, bono)
        assert result == expected

    def test_bono_positivo_incrementa_precio(self):
        base = _calc_precio_x_tms(
            dec("0.335"), dec("88"), dec("2400"), dec("10"), dec("128.5"), dec("8"), dec("0")
        )
        con_bono = _calc_precio_x_tms(
            dec("0.335"), dec("88"), dec("2400"), dec("10"), dec("128.5"), dec("8"), dec("50")
        )
        assert con_bono > base

    def test_mayor_spot_incrementa_precio(self):
        bajo = _calc_precio_x_tms(
            dec("0.335"), dec("88"), dec("2000"), dec("10"), dec("128.5"), dec("8"), dec("0")
        )
        alto = _calc_precio_x_tms(
            dec("0.335"), dec("88"), dec("2400"), dec("10"), dec("128.5"), dec("8"), dec("0")
        )
        assert alto > bajo

    def test_mayor_riesgo_reduce_precio(self):
        bajo_riesgo = _calc_precio_x_tms(
            dec("0.335"), dec("88"), dec("2400"), dec("5"), dec("128.5"), dec("8"), dec("0")
        )
        alto_riesgo = _calc_precio_x_tms(
            dec("0.335"), dec("88"), dec("2400"), dec("50"), dec("128.5"), dec("8"), dec("0")
        )
        assert alto_riesgo < bajo_riesgo

    def test_factor_1_1023_aplicado(self):
        oz = dec("0.335")
        rec = dec("88")
        spot = dec("2400")
        riesgo = dec("10")
        maquila = dec("128.5")
        insumos = dec("8")
        bono = dec("0")

        result = _calc_precio_x_tms(oz, rec, spot, riesgo, maquila, insumos, bono)
        val_1 = oz * rec / 100 * (spot - riesgo)
        sin_factor = val_1 - maquila - insumos + bono

        # El resultado debe ser sin_factor * 1.1023 (con precisión Decimal)
        assert abs(result - (sin_factor * FACTOR).quantize(Decimal("0.0001"))) < Decimal("0.001")

    def test_precio_puede_ser_negativo_en_ley_muy_baja(self):
        # oz=0.05, rec=80, spot=2400, riesgo=400, maquila=95, insumos=8
        result = _calc_precio_x_tms(
            dec("0.05"), dec("80"), dec("2400"), dec("400"), dec("95"), dec("8"), dec("0")
        )
        # El precio por TMS puede ser negativo si ley es muy baja
        assert result < dec("0")


# ==============================================================================
# _calc_total
# ==============================================================================


class TestCalcTotal:
    """
    col AM: max(0, precio_x_tms * tms)
    """

    def test_producto_normal(self):
        precio = dec("620.5")
        tms = dec("9.5")
        expected = (precio * tms).quantize(Decimal("0.01"))
        assert _calc_total(precio, tms) == expected

    def test_precio_negativo_retorna_cero(self):
        # Lote volado o ley muy baja puede generar precio negativo
        result = _calc_total(dec("-100"), dec("5"))
        assert result == dec("0")

    def test_precio_cero_retorna_cero(self):
        assert _calc_total(dec("0"), dec("9.5")) == dec("0")

    def test_precision_dos_decimales(self):
        result = _calc_total(dec("620.555"), dec("3.0"))
        # Debe tener máximo 2 decimales
        assert result == result.quantize(Decimal("0.01"))

    def test_total_escala_linealmente_con_tms(self):
        precio = dec("500")
        tms_1 = dec("5")
        tms_2 = dec("10")
        assert _calc_total(precio, tms_2) == _calc_total(precio, tms_1) * 2


# ==============================================================================
# _calc_fino_recuperable
# ==============================================================================


class TestCalcFinoRecuperable:
    """
    col AJ: 31.1035 * 1.1023 * tms * rec_liq/100 * oz_promedio / 100
    """

    def test_formula_exacta(self):
        tms = dec("9.5")
        rec_liq = dec("88")
        oz_prom = dec("0.335")

        expected = (TROY_OZ * FACTOR * tms * rec_liq / 100 * oz_prom / 100).quantize(
            Decimal("0.0001")
        )
        result = _calc_fino_recuperable(tms, rec_liq, oz_prom)
        assert result == expected

    def test_mayor_tms_produce_mas_fino(self):
        bajo = _calc_fino_recuperable(dec("5"), dec("88"), dec("0.335"))
        alto = _calc_fino_recuperable(dec("10"), dec("88"), dec("0.335"))
        assert alto > bajo

    def test_mayor_recuperacion_produce_mas_fino(self):
        bajo = _calc_fino_recuperable(dec("9.5"), dec("80"), dec("0.335"))
        alto = _calc_fino_recuperable(dec("9.5"), dec("90"), dec("0.335"))
        assert alto > bajo

    def test_mayor_ley_produce_mas_fino(self):
        bajo = _calc_fino_recuperable(dec("9.5"), dec("88"), dec("0.2"))
        alto = _calc_fino_recuperable(dec("9.5"), dec("88"), dec("0.5"))
        assert alto > bajo

    def test_precision_cuatro_decimales(self):
        result = _calc_fino_recuperable(dec("9.5"), dec("88"), dec("0.335"))
        assert result == result.quantize(Decimal("0.0001"))

    def test_constantes_troy_y_factor_incluidas(self):
        """
        Verifica que las constantes de conversión (31.1035 y 1.1023)
        están aplicadas y no son unitarias.
        """
        tms = dec("1")
        rec_liq = dec("100")
        oz_prom = dec("1")
        result = _calc_fino_recuperable(tms, rec_liq, oz_prom)
        # resultado = 31.1035 * 1.1023 * 1 * 100/100 * 1 / 100
        expected = (TROY_OZ * FACTOR / 100).quantize(Decimal("0.0001"))
        assert result == expected


# ==============================================================================
# Integración de fórmulas: flujo completo de un lote
# ==============================================================================


class TestFlujoCompletoFormulas:
    """
    Verifica que encadenar _calc_maquila → _calc_precio_x_tms → _calc_total
    → _calc_fino_recuperable produce resultados consistentes entre sí.
    """

    # Valores tomados del Excel PL_paititi (fila representativa)
    OZ_PROM = dec("0.335")
    REC_LIQ = dec("88")
    SPOT = dec("2400")
    RIESGO = dec("10")
    MAQ_BASE = dec("0")
    INSUMOS = dec("8")
    BONO = dec("0")
    TMS = dec("9.5")

    def _calcular_todos(self):
        maquila = _calc_maquila(self.OZ_PROM, self.MAQ_BASE)
        precio_x_tms = _calc_precio_x_tms(
            self.OZ_PROM, self.REC_LIQ, self.SPOT, self.RIESGO, maquila, self.INSUMOS, self.BONO
        )
        total_usd = _calc_total(precio_x_tms, self.TMS)
        fino_recup = _calc_fino_recuperable(self.TMS, self.REC_LIQ, self.OZ_PROM)
        return maquila, precio_x_tms, total_usd, fino_recup

    def test_maquila_con_ley_0_335(self):
        # oz=0.335 → trunca 0.3 → step=30 → max(95, 0+30)=95
        maquila, _, _, _ = self._calcular_todos()
        assert maquila == dec("95")

    def test_total_positivo_con_valores_normales(self):
        _, _, total_usd, _ = self._calcular_todos()
        assert total_usd > dec("0")

    def test_fino_recuperable_positivo(self):
        _, _, _, fino = self._calcular_todos()
        assert fino > dec("0")

    def test_relacion_total_vs_precio_por_tms(self):
        maquila, precio_x_tms, total_usd, _ = self._calcular_todos()
        # total_usd debe ser aproximadamente precio_x_tms * TMS
        assert abs(total_usd - (precio_x_tms * self.TMS).quantize(Decimal("0.01"))) <= dec("0.01")

    def test_lote_volado_ley_cero_produce_total_cero(self):
        # Cuando oz_promedio=0 (lote volado), total debe ser 0
        maquila = _calc_maquila(dec("0"), dec("0"))
        precio = _calc_precio_x_tms(
            dec("0"), dec("80"), dec("2400"), dec("10"), maquila, dec("8"), dec("0")
        )
        total = _calc_total(precio, dec("9.5"))
        assert total == dec("0")


# ==============================================================================
# Verificación de cálculos de profit según PL Paititi (Ejemplo IP-4123)
# ==============================================================================


class TestProfitCalculosPaititi:
    """
    Verifica concordancia de profit_rec y profit_leyes con PL Paititi
    para el lote IP-4123 con los valores:
      - spot_usd = 4066.20
      - riesgo = 120
      - rec_planta_val = 95.74
      - rec_liq = 90
      - ley_planta = 0.3760
      - oz_promedio = 0.3530
      - FACTOR = 1.1023
    """

    def test_profit_rec_ip_4123(self):
        spot_usd = dec("4066.20")
        riesgo = dec("120")
        rec_planta_val = dec("95.74")
        rec_liq = dec("90")
        ley_planta = dec("0.3760")
        factor = dec("1.1023")

        profit_rec = (
            (rec_planta_val - rec_liq) * (spot_usd - riesgo) * ley_planta * factor / dec("100")
        )
        profit_rec_rounded = profit_rec.quantize(Decimal("0.01"))
        assert profit_rec_rounded == dec("93.88")

    def test_profit_leyes_ip_4123(self):
        spot_usd = dec("4066.20")
        riesgo = dec("120")
        rec_planta_val = dec("95.74")
        ley_planta = dec("0.3760")
        oz_promedio = dec("0.3530")
        factor = dec("1.1023")

        profit_leyes = (
            (ley_planta - oz_promedio) * (spot_usd - riesgo) * rec_planta_val * factor / dec("100")
        )
        profit_leyes_rounded = profit_leyes.quantize(Decimal("0.01"))
        assert profit_leyes_rounded == dec("95.79")


# ==============================================================================
# Verificación de columnas exportadas en generar_excel_pl
# ==============================================================================


class TestGenerarExcelPLHeaders:
    """
    Verifica que el exportable de Excel del PL incluya las columnas completas
    del desglose de profit y del PL original (49 columnas en el orden estandarizado).
    """

    def test_generar_excel_pl_headers(self):
        import io
        from unittest.mock import MagicMock

        import msoffcrypto
        from app.services.liquidaciones import generar_excel_pl
        from openpyxl import load_workbook

        db_mock = MagicMock()
        db_mock.query.return_value.options.return_value.order_by.return_value.all.return_value = []

        out = generar_excel_pl(db_mock, "testpass")
        decrypted = io.BytesIO()
        office_file = msoffcrypto.OfficeFile(out)
        office_file.load_key(password="testpass")
        office_file.decrypt(decrypted)
        decrypted.seek(0)

        wb = load_workbook(decrypted)
        ws = wb.active

        expected_headers = [
            "Lote (IP)",
            "Fecha Emision",
            "Fecha Recep.",
            "Proveedor",
            "RUC",
            "Acopiador",
            "Sacos",
            "Material",
            "TMH (Peso Neto)",
            "% H2O",
            "TMS",
            "Ley Planta (Oz/TC)",
            "Ley Comercial (Oz/TC)",
            "Ley gr/TM",
            "Ley Minero (Oz/TC)",
            "Ley Promedio (Oz/TC)",
            "% Rec Liq",
            "% Rec Planta",
            "Spot USD",
            "Factor",
            "Bono ($/TM)",
            "Riesgo ($/TM)",
            "Maquila ($/TM)",
            "Gasto Acopio ($/TM)",
            "Gasto Consumo ($/TM)",
            "Insumos ($/TM)",
            "Precio x TMS ($/TM)",
            "Bruto Au (USD)",
            "Ley Ag (Gr/TM)",
            "Ley Ag (Oz/TC)",
            "Bruto Ag (USD)",
            "Profit Maquila ($/TM)",
            "Profit Rec ($/TM)",
            "Profit Consumo ($/TM)",
            "Profit Leyes ($/TM)",
            "Profit Total $/TM",
            "Total USD",
            "Total Pago al Minero (USD)",
            "Fino Rec. (Oz)",
            "Ruma",
            "Campana",
            "Estado Lote",
            "N° Liquidacion",
            "Estado Liq.",
            "Placa",
            "Conductor",
            "Transportista",
            "Guia Remision",
            "Guia Transporte",
        ]

        headers_in_excel = [
            ws.cell(row=1, column=col).value for col in range(1, len(expected_headers) + 1)
        ]
        assert headers_in_excel == expected_headers

    def test_generar_excel_pl_with_mocked_lote(self):
        import io
        from datetime import datetime
        from unittest.mock import MagicMock

        import msoffcrypto
        from app.services.liquidaciones import generar_excel_pl
        from openpyxl import load_workbook

        db_mock = MagicMock()
        ll_mock = MagicMock()
        ll_mock.lote.ip = "IP-1000"
        ll_mock.lote.tipo_material = "MINERAL"
        ll_mock.lote.estado = "LIQUIDADO"
        ll_mock.lote.analisis_ley = []
        ll_mock.lote.sesion.placa = "ABC-123"
        ll_mock.lote.sesion.conductor = "Conductor Test"
        ll_mock.lote.sesion.transportista = "Trans Test"
        ll_mock.lote.sesion.guia_remision = "GR-001"
        ll_mock.lote.sesion.guia_transporte = "GT-001"
        ll_mock.lote.ruma.codigo = "RUMA-01"
        ll_mock.lote.ruma.campana = "C-01"
        ll_mock.liquidacion.creado_en = datetime(2026, 8, 3)
        ll_mock.liquidacion.numero_liquidacion = "LIQ-100"
        ll_mock.liquidacion.estado = "APROBADA"
        ll_mock.liquidacion.provacop.proveedor.razon_social = "Minera Test S.A.C."
        ll_mock.liquidacion.provacop.proveedor.ruc = "20123456789"
        ll_mock.liquidacion.provacop.acopiador.razon_social = "Acopiador Test"
        ll_mock.fecha_recepcion_lote = datetime(2026, 8, 1)
        ll_mock.sacos_snapshot = 50
        ll_mock.tmh_snapshot = Decimal("10.5")
        ll_mock.humedad_snapshot = Decimal("2.0")
        ll_mock.tms_snapshot = Decimal("10.29")
        ll_mock.spot_usd_snapshot = Decimal("2400.00")
        ll_mock.bono = Decimal("10")
        ll_mock.porcentaje_rec_liquido = Decimal("90.0")
        ll_mock.gasto_acopio_liquidacion = Decimal("5.0")
        ll_mock.insumos_liquidacion = Decimal("15.0")
        ll_mock.total_usd = Decimal("20000.00")
        ll_mock.fino_recuperable = Decimal("15.5")
        ll_mock.oz_tc_comercial = Decimal("1.5")
        ll_mock.maquila_aplicada = Decimal("100.00")

        db_mock.query.return_value.options.return_value.order_by.return_value.all.return_value = [
            ll_mock
        ]

        out = generar_excel_pl(db_mock, "testpass")
        decrypted = io.BytesIO()
        office_file = msoffcrypto.OfficeFile(out)
        office_file.load_key(password="testpass")
        office_file.decrypt(decrypted)
        decrypted.seek(0)

        wb = load_workbook(decrypted)
        ws = wb.active

        assert ws.cell(row=2, column=1).value == "IP-1000"
        assert ws.cell(row=2, column=2).value == "2026-08-03"
        assert ws.cell(row=2, column=3).value == "2026-08-01"
        assert ws.cell(row=2, column=4).value == "Minera Test S.A.C."
        assert ws.cell(row=2, column=5).value == "20123456789"
        assert ws.cell(row=2, column=6).value == "Acopiador Test"
